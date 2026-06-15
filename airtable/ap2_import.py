#!/usr/bin/env python3
"""
AP2 Airtable Data Import — haystacked-AGV-PoC
Imports Companies → Base Models → Products → Extensions from the seed xlsx.

Linking strategy (IDs in xlsx are all placeholders):
  - Products  → Companies:    via company_id shortcode ([Balyo], [JH], …) mapped to L1 full names
  - Products  → Base Models:  by row index (L2[i] corresponds to L3[i])
  - Extensions → Base Models: by row index (L3[i] creates both the Base Model and its Extension)

Prerequisites:
  - ap2_schema.py must have run and created airtable_schema_ids.json (in the same directory)
  - export AIRTABLE_BASE_ID=appXXXXXXXXXXXXXX
  - export AIRTABLE_TOKEN=patXXXXXXXXXXXXXX

Run:
  python3 ap2_import.py --xlsx /path/to/haystacked_supplier_seed_v03.xlsx
"""

import os, sys, json, time, uuid, argparse
from pathlib import Path
import requests
import openpyxl

# ── Args ──────────────────────────────────────────────────────────────────────

parser = argparse.ArgumentParser()
parser.add_argument("--xlsx", required=True)
args = parser.parse_args()

SHEET_L1 = "Companies (L1)"
SHEET_L2 = "Products (L2)"
SHEET_L3 = "Base Model Extensions (L3)"

# ── Config ────────────────────────────────────────────────────────────────────

BASE_ID = os.environ.get("AIRTABLE_BASE_ID", "")
TOKEN   = os.environ.get("AIRTABLE_TOKEN", "")

if not BASE_ID or not TOKEN:
    sys.exit("ERROR: Set AIRTABLE_BASE_ID and AIRTABLE_TOKEN environment variables first.")

if not Path(args.xlsx).exists():
    sys.exit(f"ERROR: xlsx not found: {args.xlsx}")

schema_file = Path(__file__).parent / "airtable_schema_ids.json"
if not schema_file.exists():
    sys.exit("ERROR: airtable_schema_ids.json not found. Run ap2_schema.py first.")

with open(schema_file) as f:
    TABLE_IDS = json.load(f)["table_ids"]

HEADERS = {"Authorization": f"Bearer {TOKEN}", "Content-Type": "application/json"}
REC_URL = f"https://api.airtable.com/v0/{BASE_ID}"

# Company shortcode → L1 full company_name mapping
# These are the [placeholder] values in L2's company_id column
COMPANY_SHORTCODE = {
    "[Balyo]":  "Balyo",
    "[JH]":     "Jungheinrich AG",
    "[DS]":     "DS Automotion GmbH",
    "[STILL]":  "STILL GmbH",
    "[E80]":    "E80 Group S.p.A. (Elettric80)",
    "[KIV]":    "KIVNON Logística S.L.",
    "[MiR]":    "Mobile Industrial Robots A/S (MiR)",
    "[IW]":     "idealworks GmbH",
    "[AGX]":    "AGILOX Services GmbH",
    "[GP]":     "Geek+ (Geekplus Technology Co.)",
}

# ── API helpers ───────────────────────────────────────────────────────────────

def _call(method, url, **kw):
    r = getattr(requests, method)(url, headers=HEADERS, **kw)
    if not r.ok:
        print(f"    ERROR {r.status_code}: {r.text[:300]}")
        r.raise_for_status()
    time.sleep(0.25)
    return r.json()

def batch_create(table_key, records):
    """POST records in batches of 10. Returns list of created record dicts."""
    tid = TABLE_IDS[table_key]
    created = []
    for i in range(0, len(records), 10):
        batch = records[i:i+10]
        resp = _call("post", f"{REC_URL}/{tid}",
                     json={"records": [{"fields": r} for r in batch], "typecast": True})
        created.extend(resp["records"])
        print(f"    ... {min(i+10, len(records))}/{len(records)}")
    return created

def batch_update(table_key, updates):
    """PATCH records in batches of 10. updates = list of {id, fields}."""
    tid = TABLE_IDS[table_key]
    for i in range(0, len(updates), 10):
        batch = updates[i:i+10]
        _call("patch", f"{REC_URL}/{tid}",
              json={"records": [{"id": u["id"], "fields": u["fields"]} for u in batch], "typecast": True})

# ── Value parsers ─────────────────────────────────────────────────────────────

def parse_ms(val):
    """Pipe-separated → list. Returns None if empty."""
    if not val or str(val).strip() in ("", "nan"):
        return None
    parts = [v.strip() for v in str(val).split("|") if v.strip()]
    return parts or None

def parse_bool(val):
    s = str(val).strip().upper()
    if s == "TRUE":  return True
    if s == "FALSE": return False
    return None  # blank = unknown → omit key

def parse_int(val, allow_zero=True):
    try:
        v = int(float(str(val).strip()))
        return v if (allow_zero or v != 0) else None
    except (ValueError, TypeError):
        return None

def parse_float(val):
    try:
        v = float(str(val).strip())
        return v if v == v else None  # reject NaN
    except (ValueError, TypeError):
        return None

def parse_text(val):
    if val is None: return None
    s = str(val).strip()
    return s if s and s.lower() != "nan" else None

def parse_date(val):
    """Handles YYYY-MM-DD and YYYY-MM (→ YYYY-MM-01)."""
    s = parse_text(val)
    if not s: return None
    s = s[:10]
    if len(s) == 7:  # YYYY-MM
        s = s + "-01"
    return s if len(s) == 10 else None

def omit_none(d):
    return {k: v for k, v in d.items() if v is not None}

# ── xlsx reader ───────────────────────────────────────────────────────────────

def read_sheet(wb, name):
    if name not in wb.sheetnames:
        sys.exit(f"ERROR: Sheet '{name}' not found. Available: {wb.sheetnames}")
    ws = wb[name]
    headers = [str(c.value).strip() if c.value is not None else f"_col{i}"
               for i, c in enumerate(ws[1])]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    return rows

# ── Field lists for L3 → Extensions ──────────────────────────────────────────

EXT_BOOL = [
    "infrastructure_required", "autonomous_obstacle_bypass", "omnidirectional_movement",
    "multi_load_compatibility", "outdoor_capable", "autonomous_charging", "battery_swap_capable",
    "vda5050_compatible", "multi_fleet_capable", "manual_usage",
    "vna_capable", "forks_free_floating", "stacking_capability",
    "barcode_readers", "stock_line_scanning", "trailer_loading", "trailer_unloading", "busbar_compatible",
    "auto_hitch", "intersection_management",
    "grid_required", "rotation_capable", "rack_pin_compatible", "free_lift_open_closed_pallet",
    "task_interleaving", "onboard_ui",
    # AP0 v0.6 additions
    "ergonomic_height_adjustable", "multi_language_display", "gamification",
]

EXT_INT = [
    "length_mm", "width_mm", "operating_temp_min_c", "operating_temp_max_c",
    "operating_humidity_max_pct", "stop_accuracy_mm", "charge_time_min", "max_fleet_size",
    "lifting_height_mm", "min_total_height_mm", "min_aisle_width_mm",
    "drop_accuracy_lat_mm", "drop_accuracy_dep_mm", "drop_accuracy_angle_deg",
    "pick_req_accuracy_lat_mm", "pick_req_accuracy_dep_mm", "pick_req_accuracy_angle_deg",
    "max_trailers", "turning_radius_mm",  # auto_hitch_position_tolerance_mm mapped explicitly below (xlsx uses old name)
    "lift_height_mm", "min_ground_clearance_mm", "min_turning_radius_mm",
    "shelf_height_mm", "throughput_picks_per_hour",
    # AP0 v0.6 additions
    "min_fleet_size", "min_grid_area_m2", "concurrent_robots_per_station",
    "order_lines_per_run", "onboard_container_count",
]

EXT_FLOAT = [
    "max_payload_kg", "max_speed_ms", "max_gradient_pct", "battery_runtime_h", "towing_capacity_kg",
    "storage_density_factor",  # AP0 v0.6 addition
]

EXT_MULTI = [
    "navigation_type", "load_type", "battery_type", "safety_standard",
    "integration_capability", "station_applications", "industries_served",
    "special_fork_option", "load_detection", "guidance",
    "coupling_type", "load_transfer", "trailer_steering_technology",
    "workflow_capability", "top_module_type", "onboard_container_type", "wms_integration_native",
    # AP0 v0.6 additions
    "installation_process", "modification_process",
]

EXT_SINGLE = [
    "agv_type", "ingress_protection_rating", "cleanroom_class", "floor_flatness_req",
    "functional_safety_level", "safety_coverage", "fleet_management_system", "fleet_control_architecture",
    "fork_spread", "mast_type", "drive_type",
    "train_configuration", "route_type", "route_programming",
    "picking_mechanism", "storage_system_type", "throughput_basis",
]

# ── Phase 1: Companies ────────────────────────────────────────────────────────

def import_companies(l1_rows):
    print(f"\n=== Phase 1: Importing {len(l1_rows)} Companies ===")
    records = []
    names   = []

    for row in l1_rows:
        new_uuid = str(uuid.uuid4())
        name     = parse_text(row.get("company_name")) or "UNKNOWN"
        names.append(name)

        f = omit_none({
            "company_name":         name,
            "company_id":           new_uuid,
            "country":              parse_text(row.get("country")),
            "hq_city":              parse_text(row.get("hq_city")),
            "hq_address":           parse_text(row.get("hq_address")),
            "hq_maps_link":         parse_text(row.get("hq_maps_link")),
            "all_sites":            parse_text(row.get("all_sites")),
            "employee_count_range": parse_text(row.get("employee_count_range")),
            "fleet_size_vehicles":  parse_text(row.get("fleet_size_vehicles")),
            "founding_year":        parse_int(row.get("founding_year")),
            "website":              parse_text(row.get("website")),
            "email":                parse_text(row.get("email")),
            "phone":                parse_text(row.get("phone")),
            "contact_person":       parse_text(row.get("contact_person")),
            "certifications_generic": parse_ms(row.get("certifications_generic")),
            "languages_spoken":     parse_ms(row.get("languages_spoken")),
            "export_capable":       parse_bool(row.get("export_capable")),
            "service_coverage":     parse_ms(row.get("service_coverage")),
            "last_updated":         parse_date(row.get("last_updated")),
            # NOTES / source not imported (no source_notes field in Companies)
        })
        records.append(f)

    created = batch_create("companies", records)
    # Map: company_name → airtable record id
    name_to_id = {names[i]: created[i]["id"] for i in range(len(created))}
    print(f"    ✓ {len(created)} Companies created")
    return name_to_id

# ── Phase 2: Base Models ──────────────────────────────────────────────────────

def import_base_models(l3_rows):
    """Each L3 row = one Base Model. Returns list of airtable record ids (index-aligned with l3_rows)."""
    print(f"\n=== Phase 2: Importing {len(l3_rows)} Base Models (from L3) ===")
    records = []

    for row in l3_rows:
        new_uuid = str(uuid.uuid4())
        f = omit_none({
            "base_model_name": parse_text(row.get("model_name")) or "UNKNOWN",
            "base_model_id":   new_uuid,
            "agv_type":        parse_text(row.get("agv_type")),
            # oem_company_id not in seed → skip; oem_link_public default FALSE (unchecked)
        })
        records.append(f)

    created = batch_create("base_models", records)
    # Return index-aligned list of Airtable record IDs
    bm_ids = [rec["id"] for rec in created]
    print(f"    ✓ {len(created)} Base Models created")
    return bm_ids

# ── Phase 3: Products ─────────────────────────────────────────────────────────

def import_products(l2_rows, company_name_to_id, bm_row_ids):
    print(f"\n=== Phase 3: Importing {len(l2_rows)} Products ===")
    records = []
    link_warnings = []

    for i, row in enumerate(l2_rows):
        new_uuid = str(uuid.uuid4())

        # Resolve company link via shortcode
        shortcode = parse_text(row.get("company_id")) or ""
        co_name   = COMPANY_SHORTCODE.get(shortcode)
        co_at_id  = company_name_to_id.get(co_name) if co_name else None
        if not co_at_id:
            link_warnings.append(f"  row {i+1}: unknown company shortcode '{shortcode}'")

        # Resolve base model link by row index
        bm_at_id = bm_row_ids[i] if i < len(bm_row_ids) else None

        f = omit_none({
            "product_name":        parse_text(row.get("product_name")) or "UNKNOWN",
            "product_id":          new_uuid,
            "company_id":          [co_at_id] if co_at_id else None,
            "base_model_id":       [bm_at_id] if bm_at_id else None,
            "agv_type":            parse_text(row.get("agv_type")),
            "product_description": parse_text(row.get("product_description")),
            "reference_count":     parse_int(row.get("reference_count"), allow_zero=False),
            "min_project_value_eur": parse_int(row.get("min_project_value_eur")),
            "max_project_value_eur": parse_int(row.get("max_project_value_eur")),
            "lead_time_weeks":     parse_int(row.get("lead_time_weeks")),
            "distribution_model":  parse_text(row.get("distribution_model")),
            "is_oem_product":      parse_bool(row.get("is_oem_product")),
            "service_coverage":    parse_ms(row.get("service_coverage")),
            "active":              parse_bool(row.get("active")) if parse_bool(row.get("active")) is not None else True,
            "source_notes":        parse_text(row.get("NOTES / source")),
        })
        records.append(f)

    if link_warnings:
        print("    ⚠ Link warnings:")
        for w in link_warnings:
            print(w)

    created = batch_create("products", records)
    print(f"    ✓ {len(created)} Products created")
    return created

# ── Phase 4: Extensions ───────────────────────────────────────────────────────

def import_extensions(l3_rows, bm_row_ids):
    print(f"\n=== Phase 4: Importing {len(l3_rows)} Base Model Extensions ===")
    records = []

    for i, row in enumerate(l3_rows):
        new_uuid = str(uuid.uuid4())
        bm_at_id = bm_row_ids[i] if i < len(bm_row_ids) else None

        f = {"model_name": parse_text(row.get("model_name")) or "UNKNOWN",
             "extension_id": new_uuid}

        if bm_at_id:
            f["base_model_id"] = [bm_at_id]

        for field in EXT_BOOL:
            v = parse_bool(row.get(field))
            if v is not None: f[field] = v

        for field in EXT_INT:
            v = parse_int(row.get(field))
            if v is not None: f[field] = v

        for field in EXT_FLOAT:
            v = parse_float(row.get(field))
            if v is not None: f[field] = v

        for field in EXT_MULTI:
            v = parse_ms(row.get(field))
            if v: f[field] = v

        for field in EXT_SINGLE:
            v = parse_text(row.get(field))
            if v: f[field] = v

        for field in ("shelf_footprint_mm", "trailer_compatibility", "cart_pickup_height_range_mm"):
            v = parse_text(row.get(field))
            if v: f[field] = v

        # Renamed field: xlsx uses old name, Airtable uses AP0 canonical name
        v = parse_int(row.get("auto_hitch_pos_tol_mm"))
        if v is not None: f["auto_hitch_position_tolerance_mm"] = v

        v = parse_text(row.get("NOTES / source"))
        if v: f["source_notes"] = v

        records.append(omit_none(f))

    created = batch_create("extensions", records)
    print(f"    ✓ {len(created)} Extensions created")
    return created

# ── Validation ────────────────────────────────────────────────────────────────

def validate(n_companies, n_bm, products, extensions):
    print("\n=== Phase 5: Validation ===")
    ok = True

    def chk(cond, good, bad):
        nonlocal ok
        sym = "✓" if cond else "✗"
        print(f"    {sym} {good if cond else bad}")
        if not cond: ok = False

    chk(n_companies == 10, "10 Companies", f"Companies: {n_companies} (expected 10)")
    chk(n_bm == 52, "52 Base Models", f"Base Models: {n_bm} (expected 52)")
    chk(len(products) == 52, "52 Products", f"Products: {len(products)} (expected 52)")
    chk(len(extensions) == 52, "52 Extensions", f"Extensions: {len(extensions)} (expected 52)")

    no_co = sum(1 for p in products if not p["fields"].get("company_id"))
    no_bm = sum(1 for p in products if not p["fields"].get("base_model_id"))
    chk(no_co == 0, "All Products → Company linked", f"{no_co} Products missing company link")
    chk(no_bm == 0, "All Products → Base Model linked", f"{no_bm} Products missing base_model link")

    no_ext_bm = sum(1 for e in extensions if not e["fields"].get("base_model_id"))
    chk(no_ext_bm == 0, "All Extensions → Base Model linked", f"{no_ext_bm} Extensions missing base_model link")

    agv = {}
    for e in extensions:
        t = e["fields"].get("agv_type", "missing")
        agv[t] = agv.get(t, 0) + 1
    print(f"    agv_type distribution: {agv}")
    print(f"       (expected ~23 Forklift AGV, ~16 Tugger AGV, ~13 Mobile AMR)")

    print(f"\n{'✅ Validation PASSED' if ok else '⚠️  Validation FAILED — see ✗ lines above'}")
    return ok

# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    print(f"Loading: {args.xlsx}")
    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    l1 = read_sheet(wb, SHEET_L1)
    l2 = read_sheet(wb, SHEET_L2)
    l3 = read_sheet(wb, SHEET_L3)
    print(f"  Rows — L1: {len(l1)}, L2: {len(l2)}, L3: {len(l3)}")

    company_name_to_id = import_companies(l1)
    bm_row_ids         = import_base_models(l3)
    products           = import_products(l2, company_name_to_id, bm_row_ids)
    extensions         = import_extensions(l3, bm_row_ids)

    validate(len(company_name_to_id), len(bm_row_ids), products, extensions)

    print("""
╔═══════════════════════════════════════════════════════════╗
║  Manual checks after import (AP2 spec §9):               ║
║  • AGILOX vda5050_compatible = unchecked (NOT VDA5050)   ║
║  • MiR250 stop_accuracy_mm = 3                            ║
║  • Geek+ P-Series: grid_required + rotation_capable = ✓  ║
║  • DS AMADEUS Counter: forks_free_floating = ✓           ║
║  • STILL LTX 50 iGo: towing_capacity_kg = 5000           ║
║  • Balyo LOWY CB: forks_free_floating = ✓                ║
║  • Set view filters manually (see ap2_schema.py output)  ║
║  • Share base link with Marcus                            ║
╚═══════════════════════════════════════════════════════════╝
""")

if __name__ == "__main__":
    main()
