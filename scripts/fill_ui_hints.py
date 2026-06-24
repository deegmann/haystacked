#!/usr/bin/env python3
"""
Fill UI Hint column (col Q) in all 4 data sheets of the AP0 xlsx,
then apply format cleanup.

IMPORTANT: Opens with openpyxl (no read_only, no data_only) to preserve formulas.
Only modifies column Q (UI Hint) and formatting — no other content changes.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

XLSX_PATH = "Spec/haystacked_AP0_field_spec_v0_10.xlsx"

# ── UI Hints by field_name ──────────────────────────────────────────────────
# Key = field_name as it appears in column A of the data sheets.
# Value = German buyer-facing hint string.

UI_HINTS = {
    # ── SHARED – All AGV Types ──────────────────────────────────────────────

    # K.O. fields
    "max_payload_kg": (
        "Wie hoch ist die maximale Nutzlast pro Ladungsträger? "
        "Das AGV muss mindestens dieses Gewicht tragen können (kg)."
    ),
    "load_type": (
        "Welche Ladungsträgertypen werden transportiert? "
        "(z. B. Europalette, Gitterbox, Behälter — alle zutreffenden auswählen)"
    ),
    "agv_type": (
        "Welcher AGV-Typ ist erforderlich? (wird automatisch aus der Ausschreibung erkannt)"
    ),

    # Cond. K.O. fields
    "navigation_type": (
        "Welche Navigationsart ist vorgesehen oder erlaubt? "
        "(z. B. Natural Feature/SLAM, Magnetband, QR-Code)"
    ),
    "infrastructure_required": (
        "Ist die Installation von Infrastruktur akzeptabel? "
        "(z. B. Leitdrähte, Reflektoren, QR-Codes)"
    ),
    "outdoor_capable": (
        "Wird Außenbetrieb benötigt?"
    ),
    "multi_load_compatibility": (
        "Muss das AGV mehrere Ladungsträger gleichzeitig transportieren können?"
    ),
    "operating_temp_min_c": (
        "Wie kalt wird es am Betriebsort mindestens? (°C; z. B. –20 °C für Tiefkühlbereich)"
    ),
    "operating_temp_max_c": (
        "Wie warm wird es am Betriebsort maximal? (°C; z. B. 30 °C für Produktionshalle)"
    ),
    "operating_humidity_max_pct": (
        "Wie hoch ist die maximale relative Luftfeuchtigkeit am Betriebsort? (%)"
    ),
    "ingress_protection_rating": (
        "Welche Schutzklasse (IP-Klasse) ist am Betriebsort erforderlich? "
        "(z. B. IP54 für Spritzwasserschutz)"
    ),
    "cleanroom_class": (
        "Gibt es Reinraumanforderungen am Betriebsort? "
        "(ISO-Klasse; nur ausfüllen wenn Reinraum vorhanden)"
    ),
    "max_gradient_pct": (
        "Wie steil ist die steilste Fahrstrecke im Betriebsbereich? "
        "(% Steigung; z. B. 1,5 für eine 1,5-%-Rampe)"
    ),
    "floor_flatness_req": (
        "Gibt es besondere Anforderungen an die Bodenqualität oder -ebenheit?"
    ),
    "fleet_management_system": (
        "Gibt es ein bestehendes Flottenleitsystem, das kompatibel sein muss?"
    ),
    "vda5050_compatible": (
        "Muss das AGV den VDA-5050-Standard für Schnittstellen unterstützen?"
    ),
    "station_applications": (
        "Welche speziellen Stationsanwendungen werden benötigt? "
        "(z. B. Ladestation, Übergabestation)"
    ),
    "service_coverage": (
        "In welchen Ländern oder Regionen ist Service und Support erforderlich?"
    ),
    "country": (
        "In welchem Land befindet sich der Betriebsort?"
    ),
    "languages_spoken": (
        "Welche Sprache(n) muss der Hersteller für Support und Dokumentation beherrschen?"
    ),

    # Scoring fields
    "autonomous_obstacle_bypass": (
        "Kann das AGV Hindernissen eigenständig ausweichen ohne anzuhalten?"
    ),
    "omnidirectional_movement": (
        "Muss das AGV in alle Richtungen ohne Wenden fahren können (omnidirektional)?"
    ),
    "max_speed_ms": (
        "Wie schnell soll das AGV maximal fahren? (m/s; z. B. 1,5 m/s)"
    ),
    "stop_accuracy_mm": (
        "Wie präzise muss das AGV an Übergabe- oder Aufnahmepunkten positionieren? "
        "(mm zulässige Abweichung)"
    ),
    "battery_type": (
        "Welcher Batterietyp ist bevorzugt oder erforderlich?"
    ),
    "battery_runtime_h": (
        "Wie viele Stunden Betriebszeit pro Ladung werden mindestens benötigt? (h)"
    ),
    "charge_time_min": (
        "Wie lange darf ein vollständiger Ladevorgang maximal dauern? (min)"
    ),
    "autonomous_charging": (
        "Muss das AGV eigenständig zur Ladestation fahren und automatisch laden?"
    ),
    "battery_swap_capable": (
        "Ist ein schneller manueller Akkuwechsel statt Laden gewünscht?"
    ),
    "safety_standard": (
        "Welche Sicherheitsnormen müssen erfüllt sein? (z. B. EN ISO 3691-4)"
    ),
    "functional_safety_level": (
        "Welches funktionale Sicherheitsniveau ist gefordert? (z. B. PLd, SIL2)"
    ),
    "safety_coverage": (
        "Welche Bereiche rund um das Fahrzeug müssen durch Sicherheitssensorik abgedeckt werden?"
    ),
    "fleet_control_architecture": (
        "Welche Steuerungsarchitektur ist für die Flotte vorgesehen? "
        "(z. B. zentralisiert, dezentralisiert)"
    ),
    "max_fleet_size": (
        "Wie viele AGVs sollen gleichzeitig im Einsatz sein?"
    ),
    "multi_fleet_capable": (
        "Müssen AGVs verschiedener Hersteller im selben Bereich kooperieren können?"
    ),
    "integration_capability": (
        "Mit welchen bestehenden Systemen muss das AGV integriert werden? (WMS, ERP, MES)"
    ),
    "installation_process": (
        "Gibt es besondere Anforderungen an den Installationsprozess des AGV-Systems?"
    ),
    "modification_process": (
        "Wie einfach müssen spätere Anpassungen oder Erweiterungen möglich sein?"
    ),
    "reference_count": (
        "Anzahl der Referenzprojekte des Anbieters (wird automatisch befüllt)"
    ),
    "lead_time_weeks": (
        "Wie viele Wochen Vorlaufzeit bis zur Inbetriebnahme sind akzeptabel?"
    ),
    "employee_count_range": (
        "Mitarbeiteranzahl des Herstellers (wird automatisch befüllt)"
    ),
    "certifications_generic": (
        "Welche Zertifizierungen muss der Hersteller nachweisen? (z. B. ISO 9001)"
    ),

    # Context fields
    "length_mm": (
        "Wie lang ist das Fahrzeug? (mm; nur zur Information)"
    ),
    "width_mm": (
        "Wie breit ist das Fahrzeug? (mm; nur zur Information)"
    ),
    "min_fleet_size": (
        "Wie viele AGVs werden mindestens benötigt?"
    ),
    "typical_project_value_eur": (
        "Welches Budget steht für das AGV-Projekt zur Verfügung? (EUR)"
    ),
    "manual_usage": (
        "Soll das Fahrzeug auch manuell bedienbar sein?"
    ),
    "industries_served": (
        "In welcher Branche wird das AGV eingesetzt?"
    ),
    "distribution_model": (
        "Wie wird das AGV vertrieben bzw. geliefert?"
    ),
    "hq_city": (
        "Hauptsitz des Herstellers (wird automatisch befüllt)"
    ),
    "founding_year": (
        "Gründungsjahr des Herstellers (wird automatisch befüllt)"
    ),

    # ── Forklift AGV ────────────────────────────────────────────────────────

    # K.O. fields
    "lifting_height_mm": (
        "Bis zu welcher Höhe muss das Gerät Lasten anheben? "
        "Maximale Entnahmehöhe im Regal angeben (mm; z. B. 10.000 mm = 10 m für Hochregal)."
    ),
    "min_aisle_width_mm": (
        "Wie breit sind die schmalsten Fahrwege in Ihrer Anlage? "
        "Das AGV muss in dieser Gangbreite fahren können (mm)."
    ),

    # Cond. K.O. fields
    "min_total_height_mm": (
        "Wie hoch sind die niedrigsten Durchfahrten oder Türen im Betriebsbereich? (mm)"
    ),
    "special_fork_option": (
        "Wird eine spezielle Gabelausführung benötigt? "
        "(z. B. Seitenschieber, Klammergreifer)"
    ),
    "fork_spread": (
        "Welcher Gabelabstand ist für die Ladungsträger erforderlich? (mm)"
    ),
    "vna_capable": (
        "Ist VNA-Fähigkeit (Schmalgangbetrieb) erforderlich — oder explizit nicht gewünscht?"
    ),
    "forks_free_floating": (
        "Müssen die Gabeln seitenverschiebbar oder schwimmend gelagert sein?"
    ),
    "stacking_capability": (
        "Muss das Gerät Lasten übereinander stapeln können?"
    ),
    "barcode_readers": (
        "Sind integrierte Barcode-Scanner für die Lastidentifikation erforderlich?"
    ),
    "trailer_loading": (
        "Muss das AGV Lasten in Lkw-Trailer einfahren und beladen können?"
    ),
    "trailer_unloading": (
        "Muss das AGV Lasten aus Lkw-Trailern entnehmen können?"
    ),
    "guidance": (
        "Welches Führungskonzept soll das Regalfahrzeug im VNA-Bereich nutzen?"
    ),

    # Scoring fields
    "drop_accuracy_lat_mm": (
        "Wie präzise muss das AGV Lasten seitlich absetzen? (mm laterale Abweichung)"
    ),
    "drop_accuracy_dep_mm": (
        "Wie präzise muss das AGV Lasten in der Tiefe absetzen? (mm)"
    ),
    "drop_accuracy_angle_deg": (
        "Wie präzise muss das AGV Lasten winkelgenau absetzen? (Grad)"
    ),
    "pick_req_accuracy_lat_mm": (
        "Welche seitliche Präzision ist bei der Lastaufnahme erforderlich? (mm)"
    ),
    "pick_req_accuracy_dep_mm": (
        "Welche Tiefenpräzision ist bei der Lastaufnahme erforderlich? (mm)"
    ),
    "pick_req_accuracy_angle_deg": (
        "Welche Winkelpräzision ist bei der Lastaufnahme erforderlich? (Grad)"
    ),
    "stock_line_scanning": (
        "Muss das AGV Bestände im Regal automatisch scannen können?"
    ),
    "busbar_compatible": (
        "Ist ein Stromversorgungssystem über Schienenschleifer (Busbar) vorgesehen?"
    ),

    # Context fields
    "mast_type": (
        "Welcher Masttyp ist vorgesehen? (z. B. Duplex, Triplex, Vierstufig)"
    ),
    "drive_type": (
        "Welcher Antriebstyp ist vorgesehen? (z. B. Gegengewichtstapler, Schubmast, VNA)"
    ),
    "load_detection": (
        "Wie soll die Lastaufnahme erkannt werden?"
    ),

    # ── Tugger AGV ──────────────────────────────────────────────────────────

    # K.O. fields
    "towing_capacity_kg": (
        "Wie hoch ist das maximale Gesamtgewicht des Anhängerzuges? "
        "(kg; Summe aller Anhänger inkl. Ladung)"
    ),
    "max_trailers": (
        "Wie viele Anhänger müssen gleichzeitig gezogen werden können?"
    ),
    "coupling_type": (
        "Welche Kupplungsart verwenden die vorhandenen oder geplanten Anhänger?"
    ),
    "route_type": (
        "Welche Art von Fahrstrecke ist geplant? (feste Route oder flexible Route)"
    ),
    "turning_radius_mm": (
        "Wie viel Platz steht in Kurven zur Verfügung? "
        "(mm; der Wenderadius des Fahrzeugs muss kleiner sein)"
    ),
    "tugger_min_aisle_width_mm": (
        "Wie breit sind die schmalsten Durchfahrten im gesamten Fahrtbereich "
        "des Tuggers inkl. Anhänger? (mm)"
    ),

    # Cond. K.O. fields
    "auto_hitch": (
        "Muss der Tugger Anhänger automatisch ankuppeln können?"
    ),

    # Scoring fields
    "auto_hitch_position_tolerance_mm": (
        "Welche Positioniertoleranz ist beim automatischen Ankuppeln akzeptabel? (mm)"
    ),
    "load_transfer": (
        "Wie soll die Lastübergabe am Anhänger erfolgen?"
    ),
    "trailer_steering_technology": (
        "Welche Lenktechnologie verwenden die Anhänger? "
        "(z. B. passive Lenkung, aktive Lenkung)"
    ),
    "route_programming": (
        "Wie sollen Fahrrouten programmiert oder angepasst werden?"
    ),
    "intersection_management": (
        "Wie sollen Kreuzungssituationen zwischen mehreren AGVs gehandhabt werden?"
    ),

    # Context fields
    "train_configuration": (
        "Welche Zuganordnung ist vorgesehen? (z. B. Anhänger in Reihe, Karussellanordnung)"
    ),
    "trailer_compatibility": (
        "Welche Anhängertypen sollen verwendet werden?"
    ),

    # ── Mobile AMR ──────────────────────────────────────────────────────────

    # K.O. fields
    "lift_height_mm": (
        "Bis zu welcher Höhe muss das AMR Lasten oder Regale anheben? (mm)"
    ),
    "min_ground_clearance_mm": (
        "Wie viel Bodenfreiheit wird mindestens benötigt? "
        "(mm; z. B. für Bodenunebenheiten oder Schwellen)"
    ),
    "min_turning_radius_mm": (
        "Wie viel Platz steht für Drehbewegungen zur Verfügung? "
        "(mm; der Wendekreis des AMR muss kleiner sein)"
    ),

    # Cond. K.O. fields
    "workflow_capability": (
        "Welche Arbeitsabläufe soll das AMR unterstützen?"
    ),
    "grid_required": (
        "Ist ein physisches Gittersystem (z. B. Lagerroboter-Raster) Teil der Lösung?"
    ),
    "picking_mechanism": (
        "Welcher Mechanismus soll das AMR zum Greifen oder Kommissionieren verwenden?"
    ),
    "rack_pin_compatible": (
        "Muss das AMR mit pin-kompatiblen Lagerregalen arbeiten können?"
    ),
    "free_lift_open_closed_pallet": (
        "Muss das AMR Paletten ohne Gabelöffnungen (Vollpaletten) aufnehmen können?"
    ),
    "shelf_height_mm": (
        "Wie hoch sind die Lagerregale, die das AMR bedienen soll? (mm)"
    ),
    "shelf_footprint_mm": (
        "Wie groß ist der Grundriss der Lagereinheiten? (mm)"
    ),

    # Scoring fields
    "rotation_capable": (
        "Muss das AMR Lasten während der Fahrt drehen können?"
    ),
    "top_module_type": (
        "Welches Aufbaumodul ist erforderlich? (z. B. Förderband, Hubtisch, Greifarme)"
    ),
    "cart_pickup_height_range_mm": (
        "In welchem Höhenbereich muss das AMR Wagen oder Ladungsträger aufnehmen? (mm)"
    ),
    "pick_req_accuracy_dep_mm": (
        "Welche Tiefenpräzision ist bei der Kommissionierung erforderlich? (mm)"
    ),
    "pick_req_accuracy_angle_deg": (
        "Welche Winkelpräzision ist bei der Kommissionierung erforderlich? (Grad)"
    ),
    "drop_accuracy_lat_mm": (
        "Wie präzise muss das AMR Lasten seitlich absetzen? (mm laterale Abweichung)"
    ),
    "drop_accuracy_dep_mm": (
        "Wie präzise muss das AMR Lasten in der Tiefe absetzen? (mm)"
    ),
    "drop_accuracy_angle_deg": (
        "Wie präzise muss das AMR Lasten winkelgenau absetzen? (Grad)"
    ),
    "throughput_picks_per_hour": (
        "Wie viele Kommissionierungen pro Stunde sind mindestens erforderlich?"
    ),
    "concurrent_robots_per_station": (
        "Wie viele AMRs müssen gleichzeitig an einer Station arbeiten können?"
    ),
    "order_lines_per_run": (
        "Wie viele Auftragspositionen sollen pro Fahrt abgearbeitet werden?"
    ),
    "task_interleaving": (
        "Sollen AMRs verschiedene Aufgabentypen im Wechsel abarbeiten können?"
    ),
    "storage_density_factor": (
        "Wie wichtig ist eine hohe Lagerplatzdichte für Ihr System?"
    ),
    "ergonomic_height_adjustable": (
        "Muss die Arbeitshöhe für Mitarbeiter ergonomisch anpassbar sein?"
    ),
    "onboard_ui": (
        "Ist ein Display oder Bedienpanel am Fahrzeug erforderlich?"
    ),
    "onboard_container_type": (
        "Welcher Behältertyp soll auf dem AMR transportiert werden?"
    ),
    "onboard_container_count": (
        "Wie viele Behälter soll das AMR gleichzeitig transportieren?"
    ),
    "wms_integration_native": (
        "Ist eine direkte native Integration mit dem Warehouse-Management-System erforderlich?"
    ),

    # Context fields
    "storage_system_type": (
        "Welches Lagersystem soll das AMR bedienen?"
    ),
    "min_grid_area_m2": (
        "Wie groß ist die minimale Lagerfläche für das AMR-System? (m²)"
    ),
    "throughput_basis": (
        "Auf welcher Basis wird der Durchsatz gemessen?"
    ),
    "multi_language_display": (
        "Muss die Anzeige am Fahrzeug mehrsprachig sein?"
    ),
    "gamification": (
        "Sind spielerische Motivationselemente für Mitarbeiter gewünscht?"
    ),
}

# ── Column widths ────────────────────────────────────────────────────────────
# Col A=Field Name, B=Data Type, C=Allowed Values, D=Unit, E=Level, F=Entity,
# G=LLM Hint, H=Matching Operator, I=Scoring Weight, J=Score Function,
# K=Threshold A, L=Threshold B, M=Plaus Min, N=Plaus Max,
# O=result_card, P=Display Mode, Q=UI Hint, R=UUID

COLUMN_WIDTHS = {
    "A": 30,   # Field Name
    "B": 14,   # Data Type
    "C": 28,   # Allowed Values — wrap
    "D": 8,    # Unit
    "E": 12,   # Level
    "F": 10,   # Entity
    "G": 50,   # LLM Hint — wrap
    "H": 20,   # Matching Operator
    "I": 10,   # Scoring Weight
    "J": 16,   # Score Function
    "K": 13,   # Score Threshold A
    "L": 13,   # Score Threshold B
    "M": 13,   # Plausibility Min
    "N": 13,   # Plausibility Max
    "O": 13,   # result_card
    "P": 14,   # Display Mode
    "Q": 52,   # UI Hint — wrap
    "R": 12,   # UUID
}

WRAP_COLUMNS = {"C", "G", "Q"}  # Allowed Values, LLM Hint, UI Hint

HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FONT = Font(bold=True)
HEADER_ROW_HEIGHT = 30
DATA_ROW_HEIGHT = 15


def apply_formatting(ws):
    """Apply column widths, wrap text, and header styling to a data sheet."""
    # Set column widths
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Wrap text for designated columns (all rows)
    for col_letter in WRAP_COLUMNS:
        col_idx = ord(col_letter) - ord("A") + 1
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Header row (row 2) — bold + fill + height
    for cell in ws[2]:
        if cell.value is not None:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[2].height = HEADER_ROW_HEIGHT

    # Data rows — modest height; content height determined by wrap
    for row_idx in range(3, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = DATA_ROW_HEIGHT


DATA_SHEETS = [
    "SHARED – All AGV Types",
    "Forklift AGV",
    "Tugger AGV",
    "Mobile AMR",
]

UI_HINT_COL = 17   # column Q = index 17 (1-based)
FIELD_NAME_COL = 1 # column A


def main():
    print(f"Loading {XLSX_PATH} …")
    wb = openpyxl.load_workbook(XLSX_PATH)  # NO read_only, NO data_only

    filled_total = 0
    skipped_total = 0
    missing_total = []

    for sheet_name in DATA_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: sheet '{sheet_name}' not found — skipping")
            continue

        ws = wb[sheet_name]
        filled = 0
        skipped = 0

        for row in ws.iter_rows(min_row=3):
            field_name_cell = row[FIELD_NAME_COL - 1]   # 0-based index
            ui_hint_cell = row[UI_HINT_COL - 1]

            fn = field_name_cell.value
            if not fn or not isinstance(fn, str) or fn.startswith("──"):
                # Skip empty rows and section separator rows
                continue

            hint = UI_HINTS.get(fn)
            if hint:
                ui_hint_cell.value = hint
                filled += 1
            else:
                skipped += 1
                missing_total.append(f"{sheet_name}/{fn}")

        print(f"  {sheet_name}: {filled} filled, {skipped} no-hint")
        apply_formatting(ws)
        filled_total += filled
        skipped_total += skipped

    if missing_total:
        print(f"\n  Fields with no UI Hint defined ({len(missing_total)}):")
        for m in missing_total:
            print(f"    {m}")

    print(f"\nSaving {XLSX_PATH} …")
    wb.save(XLSX_PATH)
    print(f"Done. {filled_total} UI Hints written, {skipped_total} fields without hint.")


if __name__ == "__main__":
    main()
