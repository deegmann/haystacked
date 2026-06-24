#!/usr/bin/env python3
"""Rename AP0 columns in all DATA_SHEETS:
  'Description — what it is · where to find it · what it implies'  → 'LLM Hint'
  'Client Explanation'                                               → 'UI Hint'
"""
from pathlib import Path
import openpyxl

ROOT      = Path(__file__).parent.parent
XLSX_PATH = ROOT / "Spec" / "haystacked_AP0_field_spec_v0_10.xlsx"
DATA_SHEETS = ["SHARED – All AGV Types", "Forklift AGV", "Tugger AGV", "Mobile AMR"]

RENAMES = {
    "Description — what it is · where to find it · what it implies": "LLM Hint",
    "Client Explanation": "UI Hint",
}

wb = openpyxl.load_workbook(XLSX_PATH)

total_renamed = 0
for sheet_name in DATA_SHEETS:
    if sheet_name not in wb.sheetnames:
        print(f"  [SKIP] Sheet not found: {sheet_name}")
        continue
    ws = wb[sheet_name]

    # Find header row — first row where at least 3 cells have non-empty string values
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        non_empty = sum(1 for c in row if c and str(c).strip())
        if non_empty >= 3:
            header_row_idx = i
            break
    if header_row_idx is None:
        print(f"  [WARN] No header row found in {sheet_name}")
        continue

    header_cells = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx))[0]
    for cell in header_cells:
        if cell.value and str(cell.value).strip() in RENAMES:
            old_name = str(cell.value).strip()
            new_name = RENAMES[old_name]
            cell.value = new_name
            print(f"  Renamed '{old_name}' → '{new_name}' in {sheet_name}")
            total_renamed += 1

if total_renamed == 0:
    print("No columns renamed — already clean or column names not found.")
else:
    wb.save(XLSX_PATH)
    print(f"\nSaved {XLSX_PATH}. Total columns renamed: {total_renamed}")
