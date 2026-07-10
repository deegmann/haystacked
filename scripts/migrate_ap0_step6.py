"""Step 6 AP0 Migration: content changes (BEHAVIOR-CHANGE).

Changes applied to Spec/haystacked_AP0_field_spec_v0_10.xlsx:
- navigation_type, battery_type, fleet_management_system, max_fleet_size,
  ingress_protection_rating, floor_flatness_req: Level → Context, Operator → (clear)
- infrastructure_required → infrastructure_free: rename + LLM hint update (operator kept)
- load_type: Allowed Values updated (expanded list; normalize non-standard values in CSV separately)
- integration_capability: Level → Cond. K.O., Operator → KO_SUBSET, Allowed Values cleaned

Run: python3 scripts/migrate_ap0_step6.py
Idempotent: running twice produces the same result.
"""
import sys
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl

AP0_PATH = ROOT / "Spec" / "haystacked_AP0_field_spec_v0_10.xlsx"

# ------------------------------------------------------------------
# New content definitions
# ------------------------------------------------------------------

LOAD_TYPE_ALLOWED = (
    "None | Pallet EUR | Pallet ISO | Half-Euro | UK Pallet | Medium Euro | "
    "Tote | Plastic Bin | Bulk Bin | Roll Container | Custom Carrier"
)

INTEGRATION_CAPABILITY_ALLOWED = (
    "None | SAP | WMS | ERP | MES | REST API | MQTT | OPC-UA | Modbus"
)

INFRASTRUCTURE_FREE_HINT = (
    "Whether the AGV operates without any permanent physical infrastructure on "
    "the facility floor or walls (no reflectors, no magnetic tape, no inductive "
    "loops, no floor markings). TRUE = infrastructure-free (no permanent site "
    "modifications required); FALSE = requires permanent infrastructure. "
    "Cond. K.O.: relevant when buyer explicitly states they cannot tolerate "
    "infrastructure modifications. NULL RULE: null unless the document explicitly "
    "addresses whether permanent infrastructure modifications are required or "
    "explicitly states the system requires no site preparation."
)

# Fields to change Level → Context and clear Operator
CONTEXT_CHANGES = {
    "navigation_type",
    "battery_type",
    "fleet_management_system",
    "max_fleet_size",
    "ingress_protection_rating",
    "floor_flatness_req",
}

DATA_SHEETS = ["Global", "AGV_Shared", "AGV_Forklift", "AGV_Tugger", "AGV_AMR"]


def _get_col_idx(ws) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    header_row = rows[1] if len(rows) > 1 and rows[1][0] != rows[0][0] else rows[0]
    # Try both row 0 and row 1 as header
    for row in rows[:3]:
        if row[0] == "Field Name":
            return {h: i for i, h in enumerate(row) if h}
    return {}


def migrate_sheet(ws, sheet_name: str, changes: dict) -> int:
    """Apply changes to a single sheet. Returns count of cells modified."""
    rows = list(ws.iter_rows(min_row=1, max_row=1))
    # Check row 2 for actual header (row 1 may be description)
    header_cells = None
    for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row[0] == "Field Name":
            header_cells = list(ws.iter_rows(min_row=ri, max_row=ri))[0]
            header_row_idx = ri
            break

    if header_cells is None:
        return 0

    col_idx = {cell.value: cell.column - 1 for cell in header_cells if cell.value}

    col_field_name = col_idx.get("Field Name")
    col_level = col_idx.get("Level")
    col_operator = col_idx.get("Matching Operator")
    col_allowed = col_idx.get("Allowed Values")
    col_hint = col_idx.get("LLM Hint")

    if col_field_name is None:
        return 0

    modified = 0
    for row in ws.iter_rows(min_row=header_row_idx + 1):
        fname_cell = row[col_field_name]
        fname = fname_cell.value
        if fname not in changes:
            continue

        ch = changes[fname]

        if "level" in ch and col_level is not None:
            cell = row[col_level]
            if cell.value != ch["level"]:
                cell.value = ch["level"]
                modified += 1

        if "operator" in ch and col_operator is not None:
            cell = row[col_operator]
            new_op = ch["operator"]
            if cell.value != new_op:
                cell.value = new_op
                modified += 1

        if "allowed_values" in ch and col_allowed is not None:
            cell = row[col_allowed]
            if cell.value != ch["allowed_values"]:
                cell.value = ch["allowed_values"]
                modified += 1

        if "field_name" in ch:
            if fname_cell.value != ch["field_name"]:
                fname_cell.value = ch["field_name"]
                modified += 1

        if "hint" in ch and col_hint is not None:
            cell = row[col_hint]
            if cell.value != ch["hint"]:
                cell.value = ch["hint"]
                modified += 1

    return modified


def main():
    wb = openpyxl.load_workbook(str(AP0_PATH))

    # Build per-field change spec
    changes: dict = {}

    for fname in CONTEXT_CHANGES:
        changes[fname] = {
            "level": "Context",
            "operator": None,
        }

    # infrastructure_required → infrastructure_free
    changes["infrastructure_required"] = {
        "field_name": "infrastructure_free",
        "hint": INFRASTRUCTURE_FREE_HINT,
        # Level and Operator stay unchanged (Cond. K.O. / KO_BOOL_REQUIRED)
    }

    # load_type: update allowed values
    changes["load_type"] = {
        "allowed_values": LOAD_TYPE_ALLOWED,
    }

    # integration_capability: promote to Cond. K.O. / KO_SUBSET
    changes["integration_capability"] = {
        "level": "Cond. K.O.",
        "operator": "KO_SUBSET",
        "allowed_values": INTEGRATION_CAPABILITY_ALLOWED,
    }

    total_modified = 0
    for sheet_name in DATA_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        n = migrate_sheet(ws, sheet_name, changes)
        if n:
            print(f"  [{sheet_name}] {n} cells modified")
        total_modified += n

    wb.save(str(AP0_PATH))
    print(f"\nAP0 xlsx saved. Total cells modified: {total_modified}")

    if total_modified == 0:
        print("  (already up to date — idempotent run)")


if __name__ == "__main__":
    main()
