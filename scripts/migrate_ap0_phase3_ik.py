"""migrate_ap0_phase3_ik.py — Phase 3 IK Sprint: AP0 xlsx migration.

Adds:
  1a) variant_guide_json column to ③ Scope Registry tab
  1b) tab_name "FoodBev_Refrigeration" for the FoodBev:Refrigeration row
  1c) New FoodBev_Refrigeration worksheet with 12 IK field rows

Always creates backup first: Spec/haystacked_AP0_field_spec_v0_10_pre_phase3_backup.xlsx
"""

import json
import shutil
import uuid
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("openpyxl not installed: pip3 install openpyxl")

ROOT   = Path(__file__).parent.parent
XLSX   = ROOT / "Spec" / "haystacked_AP0_field_spec_v0_10.xlsx"
BACKUP = ROOT / "Spec" / "haystacked_AP0_field_spec_v0_10_pre_phase3_backup.xlsx"

# ── variant_guide_json value for FoodBev:Refrigeration ────────────────────────

VARIANT_GUIDE_JSON = json.dumps({
    "Process Cooling": (
        "System actively cools process media (glycol circuits, milk, fermentation tanks, "
        "process water) in food or beverage production. Temperature range typically above "
        "freezing (+2°C to +15°C). Signals: Prozesskühlung, Brauerei, Milchkühlung, "
        "Gärung, process water, glycol chiller"
    ),
    "Cold Store": (
        "System maintains a refrigerated room or warehouse for fresh product storage. "
        "Temperature range above freezing (+0°C to +8°C). Signals: Kühllager, "
        "Kühlhaus, cold store, cold room, fresh produce storage, Frischwarenlager"
    ),
    "Deep Freeze": (
        "System freezes or maintains frozen products at temperatures well below freezing "
        "(-18°C to -40°C). Includes blast freezers and deep-freeze warehouses. "
        "Signals: Tiefkühlung, Tiefkühlanlage, Schockfrostung, blast freezer, deep freeze, "
        "frozen storage, Gefrieranlage"
    ),
}, ensure_ascii=False)

# ── AGV_Shared header columns (row 2) — used as template for new tab ──────────

AGV_SHARED_HEADER = [
    "Field Name", "Data Type", "Allowed Values", "Unit", "Level", "Entity",
    "LLM Hint", "Matching Operator", "Scoring Weight", "Score Function",
    "Score Threshold A", "Score Threshold B", "Plausibility Min", "Plausibility Max",
    "result_card", "Display Mode", "UI Hint", "UUID", "Value if Null",
]

# ── 12 IK field definitions ────────────────────────────────────────────────────
# Column order matches AGV_SHARED_HEADER.
# Sentinel "@SCOPE_VARIANTS:FoodBev:Refrigeration" will be expanded by generate_all.py.

def _make_field_row(
    field_name, data_type, allowed_values, unit, level, entity, llm_hint,
    operator, plausibility_min=None, plausibility_max=None,
):
    """Return a tuple matching AGV_SHARED_HEADER column order."""
    return (
        field_name,        # Field Name
        data_type,         # Data Type
        allowed_values,    # Allowed Values
        unit,              # Unit
        level,             # Level
        entity,            # Entity
        llm_hint,          # LLM Hint
        operator,          # Matching Operator
        None,              # Scoring Weight
        None,              # Score Function
        None,              # Score Threshold A
        None,              # Score Threshold B
        plausibility_min,  # Plausibility Min
        plausibility_max,  # Plausibility Max
        None,              # result_card
        None,              # Display Mode
        None,              # UI Hint
        str(uuid.uuid4()), # UUID — freshly generated
        None,              # Value if Null
    )


IK_FIELDS = [
    _make_field_row(
        "served_categories", "Multi-Select", "@SCOPE_VARIANTS:FoodBev:Refrigeration",
        None, "K.O.", "Base Model",
        (
            "Extract ALL categories of refrigeration applications this supplier can serve. "
            "Do NOT assume from product name alone — extract only if stated explicitly. "
            "NULL RULE: if not explicitly stated, return null."
        ),
        "KO_SUBSET",
    ),
    _make_field_row(
        "cooling_capacity_kw", "Float", None,
        "kW", "K.O.", "Base Model",
        (
            "Extract the MAXIMUM total cooling capacity in kW. Use the highest stated value. "
            "Do NOT convert from tons of refrigeration unless the document provides an explicit conversion. "
            "NULL RULE: return null if no explicit kW or TR figure is found."
        ),
        "KO_IF_LT", 1, 100000,
    ),
    _make_field_row(
        "temperature_min_celsius", "Float", None,
        "°C", "K.O.", "Base Model",
        (
            "Extract the MINIMUM evaporation or supply temperature the system can achieve in "
            "°C. Use the lowest (most negative) stated temperature target. "
            "NULL RULE: return null if not stated."
        ),
        "KO_IF_GT", -60, 15,
    ),
    _make_field_row(
        "refrigerant_types", "Multi-Select", "R717|R744|R290|R134a|R32|R404A|R452A",
        None, "K.O.", "Base Model",
        (
            "Extract ALL refrigerant types supported by this system. "
            "NULL RULE: return null if not stated."
        ),
        "KO_SUBSET",
    ),
    _make_field_row(
        "certifications_ik", "Multi-Select", "PED|ATEX|EN 378|F-Gas",
        None, "K.O.", "Base Model",
        (
            "Extract ALL relevant certifications held by this system. "
            "NULL RULE: return null if not stated."
        ),
        "KO_SUBSET",
    ),
    _make_field_row(
        "cop_efficiency", "Float", None,
        None, "K.O.", "Base Model",
        (
            "Extract the COP (Coefficient of Performance) at rated conditions. "
            "Use the highest stated COP. "
            "NULL RULE: return null if not stated."
        ),
        "KO_IF_LT", 0.5, 15,
    ),
    _make_field_row(
        "cooling_medium", "Dropdown", "glycol|water|direct",
        None, "Cond. K.O.", "Base Model",
        (
            "Extract the cooling medium used: glycol circuit, water, or direct expansion. "
            "NULL RULE: return null if not stated."
        ),
        "KO_IF_NEQ",
    ),
    _make_field_row(
        "temperature_stability_k", "Float", None,
        "K", "Cond. K.O.", "Base Model",
        (
            "Extract the maximum temperature fluctuation tolerance in Kelvin. Lower is better. "
            "NULL RULE: return null if no explicit tolerance value is stated."
        ),
        "KO_IF_GT", 0.01, 10,
    ),
    _make_field_row(
        "room_volume_m3_max", "Float", None,
        "m³", "Cond. K.O.", "Base Model",
        (
            "Extract the MAXIMUM room volume in m³ this system can handle. "
            "NULL RULE: return null if not stated."
        ),
        "KO_IF_LT", 1, 1000000,
    ),
    _make_field_row(
        "humidity_control_capable", "Boolean", None,
        None, "Cond. K.O.", "Base Model",
        (
            "Does this system provide active humidity control? "
            "Only mark true if explicitly confirmed. "
            "NULL RULE: return null if not stated."
        ),
        "KO_BOOL_REQUIRED",
    ),
    _make_field_row(
        "blast_freeze_capacity_kg_h", "Float", None,
        "kg/h", "Cond. K.O.", "Base Model",
        (
            "Extract the blast freezing throughput capacity in kg/h. "
            "NULL RULE: return null if not stated."
        ),
        "KO_IF_LT", 1, 100000,
    ),
    _make_field_row(
        "pulldown_time_h", "Float", None,
        "h", "Cond. K.O.", "Base Model",
        (
            "Extract the required pulldown time in hours (time to reach target temperature). "
            "Lower is better for supplier. "
            "NULL RULE: return null if not stated."
        ),
        "KO_IF_GT", 0.1, 48,
    ),
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _find_scope_registry_header(ws):
    """Return (row_idx_1based, col_map) for ③ Scope Registry."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), 1):
        if row[0].value == "scope_id":
            col_map = {cell.value: cell.column for cell in row if cell.value is not None}
            return i, col_map
    raise SystemExit("[FEHLER] Header 'scope_id' not found in ③ Scope Registry")


# ── Step 1a + 1b: update ③ Scope Registry ─────────────────────────────────────

def _update_scope_registry(wb):
    ws = wb["③ Scope Registry"]
    header_row_idx, col_map = _find_scope_registry_header(ws)

    # Determine next available column for variant_guide_json
    all_col_indices = [v for v in col_map.values()]
    next_col = max(all_col_indices) + 1

    # 1a) Add header for variant_guide_json
    ws.cell(row=header_row_idx, column=next_col, value="variant_guide_json")
    print(f"  [③ Scope Registry] Added 'variant_guide_json' header at column {next_col}")

    # Find column indices
    c_scope_id  = col_map.get("scope_id", 1)
    c_tab_name  = col_map.get("tab_name")

    if c_tab_name is None:
        raise SystemExit("[FEHLER] 'tab_name' column not found in ③ Scope Registry header")

    # 1b) Update FoodBev:Refrigeration row
    found = False
    for row_cells in ws.iter_rows(min_row=header_row_idx + 1, values_only=False):
        scope_id_cell = row_cells[c_scope_id - 1]
        if scope_id_cell.value == "FoodBev:Refrigeration":
            # Set tab_name
            row_cells[c_tab_name - 1].value = "FoodBev_Refrigeration"
            print(f"  [③ Scope Registry] Row {scope_id_cell.row}: tab_name → 'FoodBev_Refrigeration'")
            # Set variant_guide_json
            ws.cell(row=scope_id_cell.row, column=next_col, value=VARIANT_GUIDE_JSON)
            print(f"  [③ Scope Registry] Row {scope_id_cell.row}: variant_guide_json set ({len(VARIANT_GUIDE_JSON)} chars)")
            found = True
            break

    if not found:
        raise SystemExit("[FEHLER] Row 'FoodBev:Refrigeration' not found in ③ Scope Registry")

    # Set variant_guide_json to empty string for all OTHER rows
    for row_cells in ws.iter_rows(min_row=header_row_idx + 1, values_only=False):
        scope_id_cell = row_cells[c_scope_id - 1]
        if not scope_id_cell.value:
            continue
        if scope_id_cell.value != "FoodBev:Refrigeration":
            ws.cell(row=scope_id_cell.row, column=next_col, value="")


# ── Step 1c: create FoodBev_Refrigeration worksheet ───────────────────────────

def _create_foodbev_refrigeration_tab(wb):
    tab_name = "FoodBev_Refrigeration"
    if tab_name in wb.sheetnames:
        print(f"  [FoodBev_Refrigeration] Tab already exists — overwriting")
        del wb[tab_name]

    ws = wb.create_sheet(title=tab_name)

    # Row 1: description (mirrors AGV_Shared row 1 style)
    ws.cell(row=1, column=1, value=(
        "Fields applying to Industrial Refrigeration (FoodBev domain) — "
        "written into base_model_extensions. "
        "12 IK fields: 6 KO + 6 Cond. K.O. "
        "Colour = level (red K.O. · orange Cond. K.O.)."
    ))

    # Row 2: header (identical to AGV_Shared)
    for col_idx, col_name in enumerate(AGV_SHARED_HEADER, 1):
        ws.cell(row=2, column=col_idx, value=col_name)
    print(f"  [FoodBev_Refrigeration] Header row written ({len(AGV_SHARED_HEADER)} columns)")

    # Rows 3+: IK field data
    for row_offset, field_row in enumerate(IK_FIELDS, 3):
        for col_idx, value in enumerate(field_row, 1):
            ws.cell(row=row_offset, column=col_idx, value=value)
        print(f"  [FoodBev_Refrigeration] Row {row_offset}: {field_row[0]}")

    print(f"  [FoodBev_Refrigeration] {len(IK_FIELDS)} field rows written")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("migrate_ap0_phase3_ik.py — Phase 3 IK Sprint AP0 migration")
    print("=" * 60)

    if not XLSX.exists():
        raise SystemExit(f"[FEHLER] AP0 xlsx not found: {XLSX}")

    # Backup
    print(f"\nStep 0: Backup → {BACKUP.name}")
    shutil.copy2(XLSX, BACKUP)
    print(f"  Backup created: {BACKUP}")

    print(f"\nStep 1: Loading workbook: {XLSX.name}")
    wb = openpyxl.load_workbook(str(XLSX))
    print(f"  Sheets: {wb.sheetnames}")

    print("\nStep 1a+1b: Updating ③ Scope Registry …")
    _update_scope_registry(wb)

    print("\nStep 1c: Creating FoodBev_Refrigeration tab …")
    _create_foodbev_refrigeration_tab(wb)

    print(f"\nStep 2: Saving workbook …")
    wb.save(str(XLSX))
    print(f"  Saved: {XLSX}")

    print("\nDone. Run 'python3 scripts/generate_all.py' next.")


if __name__ == "__main__":
    main()
