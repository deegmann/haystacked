#!/usr/bin/env python3
"""
Einmaliges Migrations-Skript: legt "UUID"-Spalte in allen AP0 DATA_SHEETS an
und befüllt leere Zellen mit frisch generierten UUIDs.
Idempotent: zweiter Lauf ändert nichts.
"""

import sys
from pathlib import Path
from uuid import uuid4

import openpyxl

AP0_PATH = Path(__file__).parent.parent / "Spec" / "haystacked_AP0_field_spec_v0_10.xlsx"

DATA_SHEETS = [
    "SHARED – All AGV Types",
    "Forklift AGV",
    "Tugger AGV",
    "Mobile AMR",
]


def find_header_row(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and "Field Name" in str(row):
            return i
    return None


def find_or_create_uuid_col(ws, header_row):
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        if ws.cell(header_row, c).value == "UUID":
            return c, False
    new_col = max_col + 1
    ws.cell(header_row, new_col).value = "UUID"
    return new_col, True


def process_sheet(ws, sheet_name):
    header_row = find_header_row(ws)
    if header_row is None:
        print(f"[{sheet_name}] ERROR: header row with 'Field Name' not found — skipping")
        return

    uuid_col, created = find_or_create_uuid_col(ws, header_row)
    if created:
        print(f"[{sheet_name}] UUID-Spalte neu angelegt bei col {uuid_col}")
    else:
        print(f"[{sheet_name}] UUID-Spalte bereits vorhanden bei col {uuid_col}")

    new_count = 0
    existing_count = 0

    for r in range(header_row + 1, ws.max_row + 1):
        row_key = ws.cell(r, 1).value
        if not row_key:
            continue
        if str(row_key).startswith("──"):
            continue
        cell = ws.cell(r, uuid_col)
        if cell.value:
            existing_count += 1
        else:
            cell.value = str(uuid4())
            new_count += 1

    print(f"[{sheet_name}] {new_count} UUIDs neu vergeben, {existing_count} bereits vorhanden")


def main():
    if not AP0_PATH.exists():
        print(f"ERROR: AP0 xlsx not found at {AP0_PATH}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(AP0_PATH)

    for sheet_name in DATA_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"[{sheet_name}] ERROR: sheet not found — skipping")
            continue
        process_sheet(wb[sheet_name], sheet_name)

    wb.save(AP0_PATH)
    print("Done — workbook saved.")


if __name__ == "__main__":
    main()
