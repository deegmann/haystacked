"""
generate_all.py — Single Source of Truth pipeline

Reads TWO xlsx files and generates ALL runtime config:
  AP0_[industry].xlsx        — industry-specific (fields, operators, scoring weights
                                inline, vehicle types, extraction hints in LLM Hint col)
  platform_config.xlsx       — cross-industry (NACE codes, basic extraction schema,
                                platform scope definition)

Architecture principle:
  NOTHING industry-specific is hardcoded in Python or .txt files.
  All fields, prompts, matching rules, vehicle types, scoring weights come from AP0.xlsx.
  Cross-industry config (NACE, basic extraction schema, platform scope) comes from
  platform_config.xlsx. When a new industry is added, only the xlsx files change —
  no Python code changes are needed.

Usage:
    python3 scripts/generate_all.py [--xlsx PATH] [--platform PATH] [--db PATH] [--dry-run]

Generated files:
    config/fields.json             — all field definitions keyed by UUID (AP0 SSoT)
    config/vehicle_types.json      — _VT_MAP, VNA detection, keyword fallback
    config/nace_codes.json         — NACE Prio-1 list + platform scope
    config/sqlite_schema.json      — CREATE TABLE SQL generated from ② Structure + scope tabs
    config/plausibility.json       — LLM value plausibility ranges per extraction field
    config/prompts/*.txt           — all LLM prompt files (generated, never edit manually)
    config/ap0_checksum.txt        — MD5 for startup auto-regen

tender_key is derived as "required_" + field_name (no separate "Tender JSON Key" column).
    All extraction fields are anchored in AP0 via field_name:
      required_product_type          ← product_type (SHARED, K.O.)
      required_navigation_type   ← navigation_type (SHARED, Cond. K.O.)
      required_max_payload_kg    ← max_payload_kg (SHARED, K.O.)
      required_lifting_height_mm ← lifting_height_mm (Forklift, K.O.)
      required_min_aisle_width_mm ← min_aisle_width_mm (Forklift, K.O.)
      required_outdoor_capable   ← outdoor_capable (SHARED, Cond. K.O.)
      required_operating_temp_min_c ← operating_temp_min_c (SHARED, Cond. K.O.)
      required_cleanroom_class   ← cleanroom_class (SHARED, Cond. K.O.)
      required_load_type         ← load_type (SHARED, K.O.)
      required_towing_capacity_kg ← towing_capacity_kg (Tugger, K.O.)
      required_integration_capability ← integration_capability (SHARED, Scoring)
      required_station_applications ← station_applications (SHARED, Cond. K.O.)
"""

import argparse
import hashlib
import json
import sqlite3
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed: pip3 install openpyxl")

ROOT             = Path(__file__).parent.parent
DEFAULT_XLSX     = ROOT / "Spec" / "haystacked_AP0_field_spec_v0_10.xlsx"
DEFAULT_PLATFORM = ROOT / "Spec" / "haystacked_platform_config.xlsx"
DEFAULT_DB       = ROOT / "data" / "haystacked.db"
CONFIG_DIR   = ROOT / "config"
PROMPTS_DIR  = CONFIG_DIR / "prompts"

sys.path.insert(0, str(ROOT))
from src.prompt_markers import NULL_RULE_PATTERN

LEVEL_MAP = {"K.O.": "KO", "Cond. K.O.": "COND_KO", "Scoring": "SCORING", "Context": "CONTEXT"}
VALID_OPS  = {"KO_IF_LT","KO_IF_GT","KO_IF_NEQ","KO_BOOL_REQUIRED","KO_BOOL_EXCLUSIVE","KO_SUBSET"}

# AP0 Data Type → SQLite column type mapping
SQLITE_TYPE_MAP = {
    "Boolean": "INTEGER", "Boolean (derived)": "INTEGER",
    "Integer": "INTEGER",
    "Real": "REAL", "Float": "REAL",
    "Dropdown": "TEXT", "Multi-Select": "TEXT",
    "Text": "TEXT", "Long Text": "TEXT",
    "UUID": "TEXT", "UUID (Linked)": "TEXT",
    "ISO 3166": "TEXT", "Date": "TEXT", "URL": "TEXT",
}

SQLITE_SKIP = {
    "extension_id","base_model_id","company_id","product_id","company_name","product_name",
    "product_description","base_model_name","oem_company_id","oem_link_public","website",
    "last_updated","is_oem_product","active","extra_fields",
    "pick_req_accuracy_lat_mm_amr","pick_req_accuracy_dep_mm_amr","pick_req_accuracy_angle_deg_amr",
    "drop_accuracy_lat_mm_amr","drop_accuracy_dep_mm_amr","drop_accuracy_angle_deg_amr",
    "min_project_value_eur","max_project_value_eur",
}



# ── Readers ───────────────────────────────────────────────────────────────────

def _rows(ws, min_row=1):
    return list(ws.iter_rows(min_row=min_row, values_only=True))

def _find_header(rows, key="Field Name"):
    for i, row in enumerate(rows):
        if row and row[0] == key:
            return i, {v: j for j, v in enumerate(row) if v}
    return None, {}


SCOPE_REGISTRY_TAB = "③ Scope Registry"


def read_scope_registry(wb) -> tuple:
    """Read ③ Scope Registry tab. Returns (data_sheets, shared_tab, leaf_tabs, tab_scope_map, scopes, scope_nodes).

    data_sheets: ordered list of all non-blank tab_name values from ③ Scope Registry
    shared_tab:  tab_name of the scope with parent_scope == "*" and a non-blank tab_name
    leaf_tabs:   tab_names of scopes that have no children (type-specific sheets only)
    scope_nodes: dict[scope_id → dict] with all columns per node (including 7 new Step-7 columns)

    Hard error if any non-blank tab_name is absent from the workbook.
    """
    if SCOPE_REGISTRY_TAB not in wb.sheetnames:
        sys.exit(f"[FEHLER] read_scope_registry: Tab '{SCOPE_REGISTRY_TAB}' nicht in AP0 xlsx gefunden.")

    rows = _rows(wb[SCOPE_REGISTRY_TAB])
    hi, cols = _find_header(rows, key="scope_id")
    if hi is None:
        sys.exit(f"[FEHLER] read_scope_registry: Kein Header 'scope_id' in '{SCOPE_REGISTRY_TAB}' gefunden.")

    c_scope_id          = cols.get("scope_id", 0)
    c_parent            = cols.get("parent_scope")
    c_tab_name          = cols.get("tab_name")
    c_active            = cols.get("active")
    c_canonical_name    = cols.get("canonical_name")
    c_classification    = cols.get("classification_guide")
    c_keywords          = cols.get("keywords")
    c_variants          = cols.get("variants")
    c_vna_applicable    = cols.get("vna_applicable")
    c_scoring_bucket    = cols.get("scoring_bucket")
    c_vna_hint          = cols.get("vna_hint")
    c_variant_guide_json = cols.get("variant_guide_json")

    def _str(row, c):
        return str(row[c]).strip() if c is not None and c < len(row) and row[c] else ""

    def _bool(row, c):
        v = row[c] if c is not None and c < len(row) else None
        if v is None:
            return None
        if isinstance(v, bool):
            return v
        return str(v).strip().lower() in ("true", "yes", "1")

    scopes = []  # (scope_id, parent_scope, tab_name)
    scope_nodes = {}  # scope_id → dict with all new columns
    for row in rows[hi + 1:]:
        if not row or not row[c_scope_id]:
            continue
        scope_id = str(row[c_scope_id]).strip()
        parent   = _str(row, c_parent)
        tab_name = _str(row, c_tab_name)
        active   = _str(row, c_active)
        if not active:
            continue
        scopes.append((scope_id, parent, tab_name))

        canonical_name = _str(row, c_canonical_name)
        classification = _str(row, c_classification)
        raw_keywords   = _str(row, c_keywords)
        raw_variants   = _str(row, c_variants)
        vna_applicable = _bool(row, c_vna_applicable)
        scoring_bucket = _str(row, c_scoring_bucket)
        vna_hint       = _str(row, c_vna_hint)

        node: dict = {}
        if canonical_name:
            node["canonical_name"]      = canonical_name
        if classification:
            node["classification_guide"] = classification
        if raw_keywords:
            node["scope_keywords"] = [kw.strip() for kw in raw_keywords.split(",") if kw.strip()]
        if raw_variants:
            node["scope_variants"] = [v.strip() for v in raw_variants.split(",") if v.strip()]
        if vna_applicable is not None:
            node["vna_applicable"] = vna_applicable
        if scoring_bucket:
            node["scoring_bucket"] = scoring_bucket
        if vna_hint:
            node["vna_hint"] = vna_hint
        if c_variant_guide_json is not None and c_variant_guide_json < len(row) and row[c_variant_guide_json]:
            raw_vgj = str(row[c_variant_guide_json]).strip()
            if raw_vgj:
                import json as _json
                node["variant_guides"] = _json.loads(raw_vgj)
        scope_nodes[scope_id] = node

    # Hard error: non-blank tab_name absent from workbook
    for scope_id, _, tab_name in scopes:
        if tab_name and tab_name not in wb.sheetnames:
            sys.exit(
                f"[FEHLER] read_scope_registry: tab_name '{tab_name}' (scope '{scope_id}') "
                f"nicht in AP0 xlsx vorhanden. Fehlender Tab oder Tippfehler in ③."
            )

    data_sheets = [t for _, _, t in scopes if t]

    shared_tabs = [t for sid, p, t in scopes if p == "*" and t]
    if len(shared_tabs) < 1:
        sys.exit(
            f"[FEHLER] read_scope_registry: Mindestens ein Scope mit parent_scope='*' und "
            f"nicht-leerem tab_name erwartet, gefunden: {shared_tabs}"
        )
    shared_tab = shared_tabs[0]

    parent_ids = {p for _, p, _ in scopes if p}
    leaf_tabs  = [t for sid, _, t in scopes if t and sid not in parent_ids]
    if not leaf_tabs:
        sys.exit(
            f"[FEHLER] read_scope_registry: Keine Leaf-Scopes mit tab_name gefunden. "
            f"Mindestens ein Scope ohne Kinder mit nicht-leerem tab_name wird benötigt."
        )

    tab_scope_map = {t: sid for sid, _, t in scopes if t}
    return data_sheets, shared_tab, leaf_tabs, tab_scope_map, scopes, scope_nodes


def read_field_levels(wb, data_sheets) -> dict:
    fields = {}
    for sheet in data_sheets:
        if sheet not in wb.sheetnames:
            print(f"  [WARN] Sheet missing: {sheet}")
            continue
        rows = _rows(wb[sheet])
        hi, cols = _find_header(rows)
        if hi is None:
            continue
        col_level       = cols.get("Level")
        col_op          = cols.get("Matching Operator")
        col_dtype       = cols.get("Data Type")
        col_allowed     = cols.get("Allowed Values")
        col_unit        = cols.get("Unit")
        col_display_mode = cols.get("Display Mode")
        for row in rows[hi+1:]:
            fname = row[0]
            if not fname or str(fname).startswith("──"):
                continue
            raw_level = row[col_level] if col_level is not None else None
            level = LEVEL_MAP.get(str(raw_level).strip()) if raw_level else None
            if not level:
                continue
            entry = {"level": level}
            raw_dtype = str(row[col_dtype]).strip() if col_dtype is not None and col_dtype < len(row) and row[col_dtype] else ""
            if raw_dtype:
                entry["data_type"] = raw_dtype
            if col_op is not None and col_op < len(row) and row[col_op]:
                op = str(row[col_op]).strip()
                if op in VALID_OPS:
                    entry["operator"] = op
                else:
                    print(f"  [WARN] Unknown operator '{op}' for '{fname}'")
            # Derive tender_key from field_name — no separate column needed
            entry["tender_key"] = "required_" + str(fname).strip()
            # Store allowed values for Dropdown/Multi-Select fields as a validation list
            if raw_dtype in ("Dropdown", "Multi-Select") and col_allowed is not None and col_allowed < len(row) and row[col_allowed]:
                raw_av = str(row[col_allowed]).strip()
                # Parse "A | B | C" into ["A","B","C"], skip pure unit strings
                av_list = [v.strip() for v in raw_av.split("|") if v.strip() and v.strip() not in ("…", "")]
                if av_list:
                    entry["allowed_values"] = av_list
            if col_display_mode is not None and col_display_mode < len(row) and row[col_display_mode]:
                entry["display_mode"] = str(row[col_display_mode]).strip()
            if level in ("KO","COND_KO") and "operator" not in entry:
                print(f"  [WARN] '{fname}' is {level} but has no operator")
            if fname not in fields:
                fields[str(fname)] = entry
    return fields


def read_vehicle_types(wb) -> dict:
    """Returns structured vehicle type config."""
    if "Vehicle Types" not in wb.sheetnames:
        print("  [WARN] 'Vehicle Types' sheet missing — vehicle_types.json will be empty")
        return {}
    rows = _rows(wb["Vehicle Types"], min_row=2)
    hi, cols = _find_header(rows, "LLM Output")
    if hi is None:
        print("  [WARN] No header in Vehicle Types sheet")
        return {}

    vt_map = {}         # llm_output_lower → canonical
    vna_subtypes = []   # list of llm_output_lower values
    overrides = []      # list of {regex, canonical, vna}
    keyword_map = {}    # canonical → list of keywords
    llm_guide = []      # list of {name, description, key_indicators} for prompt

    c_llm    = cols.get("LLM Output", 0)
    c_canon  = cols.get("Canonical Type", 1)
    c_vna    = cols.get("VNA Subtype", 2)
    c_kw     = cols.get("Fallback Keywords (|sep)", 3)
    c_regex  = cols.get("Text Override Regex", 4)
    c_desc   = cols.get("LLM Description", 5)
    c_ind    = cols.get("LLM Key Indicators", 6)

    for row in rows[hi+1:]:
        if not row or not row[c_llm]:
            continue
        llm_out  = str(row[c_llm]).strip()
        canon    = str(row[c_canon]).strip() if row[c_canon] else ""
        is_vna   = str(row[c_vna]).strip().lower() == "yes" if row[c_vna] else False
        kw_raw   = str(row[c_kw]).strip() if c_kw < len(row) and row[c_kw] else ""
        regex    = str(row[c_regex]).strip() if c_regex < len(row) and row[c_regex] else ""
        desc     = str(row[c_desc]).strip() if c_desc < len(row) and row[c_desc] else ""
        ind      = str(row[c_ind]).strip()  if c_ind  < len(row) and row[c_ind]  else ""

        if canon:
            vt_map[llm_out.lower()] = canon
        if is_vna:
            vna_subtypes.append(llm_out.lower())
        if regex and canon:
            overrides.append({"regex": regex, "canonical": canon, "vna": is_vna})
        if kw_raw and canon:
            kws = [k.strip() for k in kw_raw.split("|") if k.strip()]
            keyword_map.setdefault(canon, []).extend(kws)
        if desc:
            llm_guide.append({"name": llm_out, "description": desc, "key_indicators": ind})

    return {
        "vt_map":       vt_map,
        "vna_subtypes": vna_subtypes,
        "text_overrides": overrides,
        "keyword_map":  {k: list(dict.fromkeys(v)) for k, v in keyword_map.items()},
        "llm_guide":    llm_guide,
    }


def read_field_text_fallbacks(wb) -> list:
    """Read Field Fallbacks sheet → list of {tender_key, regex, value, only_if_null}."""
    if "Field Fallbacks" not in wb.sheetnames:
        return []
    rows = _rows(wb["Field Fallbacks"], min_row=2)
    hi, cols = _find_header(rows, "Tender Key")
    if hi is None:
        return []
    c_key   = cols.get("Tender Key", 0)
    c_regex = cols.get("Regex (applied to full PDF text)", 1)
    c_val   = cols.get("Fallback Value (AP0 allowed value)", 2)
    c_null  = cols.get("Only If Null", 3)
    result = []
    for row in rows[hi+1:]:
        if not row or not row[c_key]:
            continue
        tender_key = str(row[c_key]).strip()
        regex      = str(row[c_regex]).strip() if c_regex < len(row) and row[c_regex] else ""
        value      = str(row[c_val]).strip()   if c_val   < len(row) and row[c_val]   else ""
        only_null  = str(row[c_null]).strip().lower() != "no" if c_null < len(row) and row[c_null] else True
        if tender_key and regex and value:
            result.append({"tender_key": tender_key, "regex": regex, "value": value, "only_if_null": only_null})
    return result


def read_scoring_weights(wb, data_sheets) -> dict:
    """Read scoring weights + rules from AP0 xlsx.

    Reads: Scoring Weight, Score Function, Score Threshold A, Score Threshold B.
    Output format per field: {"weight": int, "rule": str, "t1": num|None, "t2": num|None}
    """
    result = {sheet: {} for sheet in data_sheets}
    sheet_map = {sheet: sheet for sheet in data_sheets}
    def _int(v):
        try: return int(v) if v is not None and v != '' else 0
        except: return 0
    def _num(v):
        if v is None or v == '': return None
        try:
            f = float(v)
            return int(f) if f == int(f) else f
        except: return None

    for sheet, bucket in sheet_map.items():
        if sheet not in wb.sheetnames: continue
        rows = _rows(wb[sheet])
        hi, cols = _find_header(rows)
        if hi is None: continue
        c_field  = cols.get("Field Name", 0)
        c_weight = cols.get("Scoring Weight")
        c_rule   = cols.get("Score Function")
        c_t1     = cols.get("Score Threshold A")
        c_t2     = cols.get("Score Threshold B")
        if c_weight is None: continue
        for row in rows[hi+1:]:
            if not row or not row[c_field]: continue
            fname = str(row[c_field]).strip()
            if fname.startswith("──"): continue
            w = _int(row[c_weight] if c_weight < len(row) else None)
            if not w: continue
            rule = (str(row[c_rule]).strip()
                    if c_rule is not None and c_rule < len(row) and row[c_rule] else "bool")
            t1   = _num(row[c_t1] if c_t1 is not None and c_t1 < len(row) else None)
            t2   = _num(row[c_t2] if c_t2 is not None and c_t2 < len(row) else None)
            entry: dict = {"weight": w, "rule": rule}
            if t1 is not None: entry["t1"] = t1
            if t2 is not None: entry["t2"] = t2
            result[bucket][fname] = entry
    return result


def read_extraction_schema(wb, data_sheets, tab_scope_map: dict = None) -> list:
    """Build extraction schema from AP0 field sheets.

    tender_key is derived as "required_" + field_name.
    LLM extraction hint comes from the "LLM Hint" column.
    """
    if tab_scope_map is None:
        tab_scope_map = {}
    schema = []
    seen = set()

    for sheet in data_sheets:
        if sheet not in wb.sheetnames: continue
        rows = _rows(wb[sheet])
        hi, cols = _find_header(rows)
        if hi is None: continue
        c_field   = cols.get("Field Name", 0)
        c_hint    = cols.get("LLM Hint")
        c_allowed = cols.get("Allowed Values")

        for row in rows[hi+1:]:
            if not row or not row[c_field]: continue
            fname = str(row[c_field]).strip()
            if fname.startswith("──"): continue
            jk = "required_" + fname
            if jk in seen: continue
            hint = str(row[c_hint]).strip() if c_hint is not None and c_hint < len(row) and row[c_hint] else ""
            if not hint:
                continue  # fields without LLM Hint are not extracted by the LLM
            seen.add(jk)
            entry = {"key": jk, "db_field": fname, "mandatory": False, "hint": hint, "scope": tab_scope_map.get(sheet, sheet)}
            raw_av = str(row[c_allowed]).strip() if c_allowed is not None and c_allowed < len(row) and row[c_allowed] else ""
            if raw_av.startswith("@SCOPE_VARIANTS:"):
                entry["variant_scope"] = raw_av.split(":", 1)[1].strip()
            schema.append(entry)

    return schema


STRUCTURE_TAB = "② Structure"


def read_structure_registry(wb) -> dict:
    """Read ② Structure tab. Returns structural columns per table.

    Returns dict keyed by table name. Each value is a list of:
      {"column": str, "sqlite_type": str, "role": str,
       "references": str, "nullable": bool, "notes": str}

    Roles: PK, FK, ADMIN, PLAIN, DERIVED.
    Business fields (KO/COND_KO/SCORING/CONTEXT) are NOT listed here —
    they live in scope tabs and are appended by read_sqlite_schema via fields.json.
    """
    if STRUCTURE_TAB not in wb.sheetnames:
        sys.exit(f"[FEHLER] read_structure_registry: Tab '{STRUCTURE_TAB}' nicht in AP0 xlsx.")

    rows = _rows(wb[STRUCTURE_TAB])
    hi, cols = _find_header(rows, key="table")
    if hi is None:
        sys.exit(f"[FEHLER] read_structure_registry: Kein Header 'table' in '{STRUCTURE_TAB}'.")

    c_table    = cols.get("table", 0)
    c_column   = cols.get("column")
    c_type     = cols.get("sqlite_type")
    c_role     = cols.get("role")
    c_refs     = cols.get("references")
    c_nullable = cols.get("nullable")
    c_notes    = cols.get("notes")

    result: dict = {}
    for row in rows[hi + 1:]:
        if not row or not row[c_table]:
            continue
        table  = str(row[c_table]).strip()
        column = str(row[c_column]).strip() if c_column is not None and c_column < len(row) and row[c_column] else ""
        if not column:
            continue
        sql_type = str(row[c_type]).strip()   if c_type     is not None and c_type     < len(row) and row[c_type]     else "TEXT"
        role     = str(row[c_role]).strip()   if c_role     is not None and c_role     < len(row) and row[c_role]     else "PLAIN"
        refs     = str(row[c_refs]).strip()   if c_refs     is not None and c_refs     < len(row) and row[c_refs]     else ""
        nullable = bool(row[c_nullable])      if c_nullable is not None and c_nullable < len(row) and row[c_nullable] else False
        notes    = str(row[c_notes]).strip()  if c_notes    is not None and c_notes    < len(row) and row[c_notes]    else ""

        result.setdefault(table, []).append({
            "column": column, "sqlite_type": sql_type, "role": role,
            "references": refs, "nullable": nullable, "notes": notes,
        })

    if not result:
        sys.exit(f"[FEHLER] read_structure_registry: Keine Zeilen in '{STRUCTURE_TAB}' gefunden.")

    return result


def _entity_fields_from_sheet(ws, entity: str) -> list:
    """Return ordered list of field_names for a given entity from a scope tab worksheet."""
    rows = _rows(ws)
    hi, cols = _find_header(rows)
    if hi is None:
        return []
    c_entity = cols.get("Entity")
    if c_entity is None:
        return []
    result = []
    seen = set()
    for row in rows[hi + 1:]:
        if not row or not row[0]:
            continue
        fname = str(row[0]).strip()
        if fname.startswith("──") or fname in seen:
            continue
        ent = str(row[c_entity]).strip() if c_entity < len(row) and row[c_entity] else ""
        if ent == entity:
            seen.add(fname)
            result.append(fname)
    return result


def read_sqlite_schema(wb, data_sheets) -> dict:
    """Build sqlite_schema.json from ② Structure (structural cols) + scope tabs (business cols).

    Structural columns (PK/FK/ADMIN/PLAIN/DERIVED) come from ② Structure.
    Business field columns are appended per entity from the scope tabs (data_sheets loop).
    Returns the same keys as before: companies, products, base_models,
    base_model_extensions SQL strings + column lists + field type lists.
    """
    struct = read_structure_registry(wb)

    def _col_sql(col_entry) -> str:
        name     = col_entry["column"]
        sql_type = col_entry["sqlite_type"]
        role     = col_entry["role"]
        nullable = col_entry["nullable"]

        if role == "PK":
            constraint = " PRIMARY KEY"
        elif role == "DERIVED":
            constraint = ""   # derived: never NOT NULL in INSERT
        elif not nullable:
            constraint = " NOT NULL"
        else:
            constraint = ""
        return f"    {name:<45} {sql_type}{constraint}"

    def _fk_sql(refs: str, col: str) -> str:
        return f"    FOREIGN KEY ({col}) REFERENCES {refs}"

    # ── companies ──────────────────────────────────────────────────────────────
    co_struct = struct.get("companies", [])
    co_lines  = ["CREATE TABLE IF NOT EXISTS companies ("]
    co_fk     = []
    for e in co_struct:
        co_lines.append(_col_sql(e) + ",")
        if e["role"] == "FK":
            co_fk.append(_fk_sql(e["references"], e["column"]))
    # Append business field columns (entity=Company) from scope tabs
    for sheet in data_sheets:
        if sheet not in wb.sheetnames:
            continue
        rows = _rows(wb[sheet])
        hi, cols = _find_header(rows)
        if hi is None:
            continue
        c_entity = cols.get("Entity")
        c_dt     = cols.get("Data Type")
        if c_entity is None or c_dt is None:
            continue
        for row in rows[hi + 1:]:
            if not row or not row[0]:
                continue
            fname = str(row[0]).strip()
            if fname.startswith("──"):
                continue
            entity = str(row[c_entity]).strip() if c_entity < len(row) and row[c_entity] else ""
            if entity != "Company":
                continue
            raw_dt = str(row[c_dt]).strip() if c_dt < len(row) and row[c_dt] else "Text"
            sql_t  = SQLITE_TYPE_MAP.get(raw_dt, "TEXT")
            # Check not already added (dedup)
            if any(e["column"] == fname for e in co_struct):
                continue
            co_lines.append(f"    {fname:<45} {sql_t},")
    for fk in co_fk:
        co_lines.append(fk + ",")
    co_lines[-1] = co_lines[-1].rstrip(",")
    co_lines.append(");")
    companies_sql = "\n".join(co_lines)

    # ── products ───────────────────────────────────────────────────────────────
    pr_struct = struct.get("products", [])
    pr_lines  = ["CREATE TABLE IF NOT EXISTS products ("]
    pr_fk     = []
    pr_struct_names = {e["column"] for e in pr_struct}
    for e in pr_struct:
        pr_lines.append(_col_sql(e) + ",")
        if e["role"] == "FK":
            pr_fk.append(_fk_sql(e["references"], e["column"]))
    # Append business field columns (entity=Product)
    pr_seen = set(pr_struct_names)
    for sheet in data_sheets:
        if sheet not in wb.sheetnames:
            continue
        rows = _rows(wb[sheet])
        hi, cols = _find_header(rows)
        if hi is None:
            continue
        c_entity = cols.get("Entity")
        c_dt     = cols.get("Data Type")
        if c_entity is None or c_dt is None:
            continue
        for row in rows[hi + 1:]:
            if not row or not row[0]:
                continue
            fname = str(row[0]).strip()
            if fname.startswith("──") or fname in pr_seen:
                continue
            entity = str(row[c_entity]).strip() if c_entity < len(row) and row[c_entity] else ""
            if entity != "Product":
                continue
            pr_seen.add(fname)
            raw_dt = str(row[c_dt]).strip() if c_dt < len(row) and row[c_dt] else "Text"
            sql_t  = SQLITE_TYPE_MAP.get(raw_dt, "TEXT")
            pr_lines.append(f"    {fname:<45} {sql_t},")
    for fk in pr_fk:
        pr_lines.append(fk + ",")
    pr_lines[-1] = pr_lines[-1].rstrip(",")
    pr_lines.append(");")
    products_sql = "\n".join(pr_lines)

    # ── base_models ────────────────────────────────────────────────────────────
    bm_struct = struct.get("base_models", [])
    bm_lines  = ["CREATE TABLE IF NOT EXISTS base_models ("]
    bm_fk     = []
    for e in bm_struct:
        bm_lines.append(_col_sql(e) + ",")
        if e["role"] == "FK":
            bm_fk.append(_fk_sql(e["references"], e["column"]))
    for fk in bm_fk:
        bm_lines.append(fk + ",")
    bm_lines[-1] = bm_lines[-1].rstrip(",")
    bm_lines.append(");")
    base_models_sql = "\n".join(bm_lines)

    # ── base_model_extensions: structural from ② + business fields from scope tabs ──
    ext_struct       = struct.get("base_model_extensions", [])
    ext_struct_names = {e["column"] for e in ext_struct}

    # Pass 1: collect global type metadata for sync_airtable.py coercion — all entities
    bool_fields        = []
    int_fields         = []
    float_fields       = []
    multiselect_fields = []
    _type_seen = set()  # dedup by field_name

    for sheet in data_sheets:
        if sheet not in wb.sheetnames:
            continue
        rows = _rows(wb[sheet])
        hi, cols = _find_header(rows)
        if hi is None:
            continue
        c_dt = cols.get("Data Type")
        if c_dt is None:
            continue
        for row in rows[hi + 1:]:
            if not row or not row[0]:
                continue
            fname = str(row[0]).strip()
            if fname.startswith("──") or fname in _type_seen:
                continue
            _type_seen.add(fname)
            raw_dt = str(row[c_dt]).strip() if c_dt < len(row) and row[c_dt] else "Text"
            if raw_dt in ("Boolean", "Boolean (derived)"):
                bool_fields.append(fname)
            elif raw_dt == "Integer":
                int_fields.append(fname)
            elif raw_dt in ("Real", "Float"):
                float_fields.append(fname)
            elif raw_dt == "Multi-Select":
                multiselect_fields.append(fname)

    # Pass 2: collect base_model_extensions business columns — Base Model entity only
    ext_fields  = []   # (name, sqlite_type)
    ext_seen    = set(ext_struct_names)  # dedup guard, init with structural columns

    for sheet in data_sheets:
        if sheet not in wb.sheetnames:
            continue
        rows = _rows(wb[sheet])
        hi, cols = _find_header(rows)
        if hi is None:
            continue
        c_dt     = cols.get("Data Type")
        c_entity = cols.get("Entity")
        if c_dt is None:
            continue
        for row in rows[hi + 1:]:
            if not row or not row[0]:
                continue
            fname = str(row[0]).strip()
            if fname.startswith("──") or fname in ext_seen:
                continue
            # Only Base Model fields belong in base_model_extensions
            entity = str(row[c_entity]).strip() if c_entity is not None and c_entity < len(row) and row[c_entity] else ""
            if entity and entity != "Base Model":
                continue
            ext_seen.add(fname)
            raw_dt = str(row[c_dt]).strip() if c_dt < len(row) and row[c_dt] else "Text"
            sql_t  = SQLITE_TYPE_MAP.get(raw_dt, "TEXT")
            ext_fields.append((fname, sql_t))

    ext_lines = ["CREATE TABLE IF NOT EXISTS base_model_extensions ("]
    for e in ext_struct:
        ext_lines.append(_col_sql(e) + ",")
    for fname, sql_t in ext_fields:
        ext_lines.append(f"    {fname:<45} {sql_t},")
    # FK clause for base_model_id (from ② struct)
    for e in ext_struct:
        if e["role"] == "FK":
            ext_lines.append(_fk_sql(e["references"], e["column"]) + ",")
    ext_lines[-1] = ext_lines[-1].rstrip(",")
    ext_lines.append(");")
    extensions_sql = "\n".join(ext_lines)

    # ── Column lists for sync_airtable.py ──────────────────────────────────────
    companies_columns  = [e["column"] for e in co_struct] + [
        fname for sheet in data_sheets
        if sheet in wb.sheetnames
        for fname in _entity_fields_from_sheet(wb[sheet], "Company")
        if fname not in {e["column"] for e in co_struct}
    ]
    products_columns   = [e["column"] for e in pr_struct] + [
        fname for sheet in data_sheets
        if sheet in wb.sheetnames
        for fname in _entity_fields_from_sheet(wb[sheet], "Product")
        if fname not in pr_struct_names
    ]
    extensions_columns = (
        [e["column"] for e in ext_struct]
        + [f for f, _ in ext_fields]
    )

    # Step 5 invariant: no Company/Product fields may appear in extensions
    _denorm_check = (set(extensions_columns) - set(e["column"] for e in ext_struct)) & (set(companies_columns) | set(products_columns))
    if _denorm_check:
        sys.exit(f"[FEHLER] read_sqlite_schema: Company/Product fields still in extensions: {_denorm_check} — check Entity column in AP0 scope tabs")

    return {
        "companies":                  companies_sql,
        "products":                   products_sql,
        "base_models":                base_models_sql,
        "base_model_extensions":      extensions_sql,
        "bool_fields":                sorted(bool_fields),
        "int_fields":                 sorted(int_fields),
        "float_fields":               sorted(float_fields),
        "multiselect_fields":         sorted(multiselect_fields),
        "companies_columns":          companies_columns,
        "products_columns":           products_columns,
        "extensions_columns":         extensions_columns,
    }


def read_plausibility(wb, data_sheets) -> dict:
    """Read plausibility ranges from AP0 xlsx — tender_key → {min, max, unit, label}.

    Unit conversion rules are read separately from haystacked_platform_config.xlsx
    via read_unit_conversions(), then combined in build_plausibility_config().
    """
    plausibility: dict = {}

    for sheet_name in data_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        rows = _rows(wb[sheet_name])
        hi, cols = _find_header(rows)
        if hi is None:
            continue

        col_pmin  = cols.get("Plausibility Min")
        col_pmax  = cols.get("Plausibility Max")
        col_unit  = cols.get("Unit")

        if None in (col_pmin, col_pmax, col_unit):
            print(f"  [WARN] Plausibility columns missing in {sheet_name} — run generate_all.py after adding them to xlsx")
            continue

        for row in rows[hi + 1:]:
            fname = row[0]
            if not fname or str(fname).startswith("──"):
                continue
            tkey = "required_" + str(fname).strip()
            pmin = row[col_pmin] if col_pmin < len(row) else None
            pmax = row[col_pmax] if col_pmax < len(row) else None
            unit = row[col_unit] if col_unit < len(row) else None

            if pmin is not None and pmax is not None:
                if tkey not in plausibility:  # first occurrence wins (SHARED before type-specific)
                    plausibility[tkey] = {
                        "min":   float(pmin),
                        "max":   float(pmax),
                        "unit":  str(unit) if unit else "",
                        "label": str(fname),
                    }

    return plausibility


def read_unit_conversions(wb_platform) -> dict:
    """Read unit conversion rules from haystacked_platform_config.xlsx — Unit Conversions sheet.

    Cross-industry rules (e.g. m ↔ mm) live in the platform config, not in the AGV AP0.
    Returns: tender_unit → {llm_alias, factor, threshold}
    """
    unit_conversions: dict = {}
    if "Unit Conversions" not in wb_platform.sheetnames:
        print("  [WARN] Unit Conversions sheet missing from platform config")
        return unit_conversions
    uc_rows = _rows(wb_platform["Unit Conversions"], min_row=3)  # row 1 = description, row 2 = header
    for row in uc_rows:
        if row and row[0] and row[1]:
            unit_conversions[str(row[0])] = {
                "llm_alias": str(row[1]),
                "factor":    float(row[2]),
                "threshold": float(row[3]),
            }
    return unit_conversions


def build_plausibility_config(plausibility: dict, unit_conversions: dict) -> dict:
    """Combine plausibility ranges with unit conversion rules into config/plausibility.json format.

    Conversion block is added when the field's Unit has a matching rule in the
    Unit Conversions sheet — no field-specific flags, no hardcoded factors.
    """
    result = {}
    for key, data in plausibility.items():
        entry = dict(data)
        conv = unit_conversions.get(data.get("unit"))
        entry["conversion"] = {
            "llm_alias": conv["llm_alias"],
            "factor":    conv["factor"],
            "threshold": conv["threshold"],
        } if conv else None
        result[key] = entry
    return result


def read_platform(wb, platform_path: Path) -> dict:
    """Read platform_config.xlsx — cross-industry: NACE, basic schema, scope.

    Accepts an already-opened workbook (wb) to avoid double-loading.
    platform_path is used only for the not-found warning message.
    """
    if wb is None:
        print(f"  [WARN] platform_config not found: {platform_path}")
        return {"scope_in": "", "scope_out": "", "codes": [], "basic_schema": []}

    # Platform scope
    scope_in = scope_out = ""
    if "Platform Scope" in wb.sheetnames:
        for row in _rows(wb["Platform Scope"]):
            if not row: continue
            if row[0] and "In Scope" in str(row[0]):
                scope_in = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if row[0] and "Out of Scope" in str(row[0]):
                scope_out = str(row[1]).strip() if len(row) > 1 and row[1] else ""

    # NACE codes
    codes = []
    if "NACE Codes" in wb.sheetnames:
        header_found = False
        for row in _rows(wb["NACE Codes"]):
            if not row: continue
            if row[0] == "NACE Code": header_found = True; continue
            if not header_found: continue
            code = str(row[0]).strip() if row[0] else ""
            name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            prio = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            hint = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            if code and name and "Prio 1" in prio:
                entry = f"{code}: {name}"
                if hint and hint != "nan": entry += f" | {hint[:70]}"
                codes.append(entry)

    # Basic extraction schema
    basic_schema = []
    if "Basic Extraction Schema" in wb.sheetnames:
        rows = _rows(wb["Basic Extraction Schema"], min_row=2)
        hi, cols = _find_header(rows, "JSON Key")
        if hi is not None:
            c_key  = cols.get("JSON Key", 0)
            c_type = cols.get("Data Type", 1)
            c_mand = cols.get("Mandatory?", 2)
            c_hint = cols.get("LLM Extraction Hint", 3)
            c_def  = cols.get("JSON Default", 4)
            c_cf   = cols.get("Contact Fallback")
            for row in rows[hi+1:]:
                if not row or not row[c_key]: continue
                key  = str(row[c_key]).strip()
                hint = str(row[c_hint]).strip() if c_hint < len(row) and row[c_hint] else ""
                mand = str(row[c_mand]).strip().lower() == "yes" if c_mand < len(row) and row[c_mand] else False
                dflt = str(row[c_def]).strip() if c_def < len(row) and row[c_def] else "null"
                cf_raw = str(row[c_cf]).strip().lower() if c_cf is not None and c_cf < len(row) and row[c_cf] else None
                contact_fallback = cf_raw if cf_raw in ("trigger", "target") else None
                basic_schema.append({
                    "key": key, "mandatory": mand, "hint": hint, "default": dflt,
                    "contact_fallback": contact_fallback,
                })

    return {"scope_in": scope_in, "scope_out": scope_out, "codes": codes, "basic_schema": basic_schema}


# ── Prompt builders ───────────────────────────────────────────────────────────

def _render_leaf_scope_lines(scope: dict, indent: str = "  ") -> list:
    """Render a leaf scope node as prompt guide lines.

    If the scope has variant_guides (sub-variant aliases), output the canonical
    description first, then each alias as an indented sub-bullet.
    Otherwise output a single line: * "canonical_name" → classification_guide.
    """
    cname = scope["canonical_name"]
    guide = scope.get("classification_guide", "")
    leaf_variants = scope.get("variant_guides", {})
    result = [f'{indent}* "{cname}" → {guide}']
    for alias, alias_desc in leaf_variants.items():
        result.append(f'{indent}  * "{alias}" → {alias_desc}')
    return result


def build_vehicle_type_template(vehicle_types: dict, scope_nodes: dict = None) -> str:
    """Pass 4a template — classify vehicle type and VNA flag only."""
    lines = [
        "Classify the required AGV/AMR vehicle type from this tender document.",
        "Output ONLY the JSON object shown below — nothing else.",
        "",
    ]
    if scope_nodes:
        leaf_scopes = [n for n in scope_nodes.values() if n.get("canonical_name") and n.get("classification_guide")]
        if leaf_scopes:
            lines.append("Vehicle type classification guide:")
            for scope in leaf_scopes:
                lines.extend(_render_leaf_scope_lines(scope))
            lines.append("  Key: PAYLOAD AND LOAD TYPE determine the vehicle — not the environment alone.")
            lines.append("")
            lines.append("  THINK STEP BY STEP:")
            names = [n["canonical_name"] for n in leaf_scopes]
            lines.append(f"  IMPORTANT: Only {len(names)} values are valid: {', '.join(repr(n) for n in names)}.")
            lines.append("  Sub-variants (Counterbalanced, Reach Truck, VNA) are NOT valid outputs — they are internal properties, not types.")
            lines.append("  (1) Towing / tugger / milk run / trailer train / Routenzug? → required_product_type='Tugger AGV'.")
            lines.append("  (2) Light load (<1000 kg) + flexible SLAM navigation + no standard floor-pallet pickup? → required_product_type='Mobile AMR'.")
            lines.append("  (3) Everything else (pallets, IBCs, forks required, racking, heavy load) → required_product_type='Forklift AGV'.")
            lines.append("      Counterbalanced, Reach Truck, AND VNA are all 'Forklift AGV'.")
            lines.append("")
    else:
        guide = vehicle_types.get("llm_guide", [])
        if guide:
            lines.append("Vehicle type classification guide:")
            seen_names = set()
            for vt in guide:
                if vt["name"] in seen_names: continue
                seen_names.add(vt["name"])
                line = f'  * "{vt["name"]}" → {vt["description"]}'
                if vt.get("key_indicators"):
                    line += f'. Signals: {vt["key_indicators"]}'
                lines.append(line)
            lines.append("  Key: PAYLOAD AND LOAD TYPE determine the vehicle — not the environment alone.")
            lines.append("")
            lines.append("  THINK STEP BY STEP:")
            lines.append("  IMPORTANT: Only three values are valid: 'Forklift AGV', 'Tugger AGV', 'Mobile AMR'.")
            lines.append("  Sub-variants (Counterbalanced, Reach Truck, VNA) are NOT valid outputs — they are internal properties, not types.")
            lines.append("  (1) Towing / tugger / milk run / trailer train / Routenzug? → required_product_type='Tugger AGV'.")
            lines.append("  (2) Light load (<1000 kg) + flexible SLAM navigation + no standard floor-pallet pickup? → required_product_type='Mobile AMR'.")
            lines.append("  (3) Everything else (pallets, IBCs, forks required, racking, heavy load) → required_product_type='Forklift AGV'.")
            lines.append("      Counterbalanced, Reach Truck, AND VNA are all 'Forklift AGV'.")
            lines.append("")
    lines += [
        "Fields:",
        "- required_product_type: MANDATORY — exactly one of: 'Forklift AGV', 'Tugger AGV', 'Mobile AMR'.",
        "",
        "DOCUMENT:",
        "{text}",
        "",
        "JSON:",
        '{"required_product_type":null}',
    ]
    return "\n".join(lines)


def build_scope_classification_template(scope_nodes: dict, domain_scope_id: str = None) -> str:
    """Generate a scope classification template from ③ Scope Registry scope nodes.

    domain_scope_id=None  → combined template (all domains, backward compat).
    domain_scope_id=<sid> → domain-specific template (leaves of that domain only).
    Returns the template content string; caller is responsible for writing to disk.
    """
    lines = [
        "Classify the required product sub-type from this tender document.",
        "Output ONLY the JSON object shown below — nothing else.",
        "",
    ]
    leaf_scopes = [
        node for scope_id, node in scope_nodes.items()
        if node.get("canonical_name") and node.get("classification_guide")
        and (domain_scope_id is None
             or scope_id == domain_scope_id
             or scope_id.startswith(domain_scope_id + ":"))
    ]
    names = [n["canonical_name"] for n in leaf_scopes]
    has_vna_types = False  # initialised before conditional; set in elif branch if needed
    # Check if this is a single-leaf domain with per-variant guides (e.g. IK)
    dom_node = scope_nodes.get(domain_scope_id, {}) if domain_scope_id else {}
    dom_variant_guides = dom_node.get("variant_guides", {})
    dom_scope_variants = dom_node.get("scope_variants", [])
    if dom_scope_variants and not dom_variant_guides:
        raise ValueError(
            f"Scope node '{domain_scope_id}' has non-empty scope_variants "
            f"({dom_scope_variants}) but no variant_guides — cannot build classification template."
        )

    if dom_variant_guides:
        # Single-leaf domain (e.g. FoodBev:Refrigeration): present scope_variants as valid outputs
        names = list(dom_variant_guides.keys())
        lines.append("Product sub-type classification guide:")
        for variant, guide in dom_variant_guides.items():
            lines.append(f'  * "{variant}" → {guide}')
        lines.append(f"  IMPORTANT: Only {len(names)} values are valid: {', '.join(repr(n) for n in names)}.")
    elif leaf_scopes:
        has_vna_types = any(n.get("vna_applicable") for n in leaf_scopes)
        lines.append("Product sub-type classification guide:")
        for scope in leaf_scopes:
            lines.extend(_render_leaf_scope_lines(scope))
        if has_vna_types:
            lines.append("  Key: PAYLOAD AND LOAD TYPE determine the vehicle — not the environment alone.")
            lines.append("")
            lines.append("  THINK STEP BY STEP:")
        lines.append(f"  IMPORTANT: Only {len(names)} values are valid: {', '.join(repr(n) for n in names)}.")
        if has_vna_types:
            lines.append("  Sub-variants (Counterbalanced, Reach Truck, VNA) are NOT valid outputs — they are internal properties, not types.")
            lines.append("  (1) Towing / tugger / milk run / trailer train / Routenzug? → required_product_type='Tugger AGV'.")
            lines.append("  (2) Light load (<1000 kg) + flexible SLAM navigation + no standard floor-pallet pickup? → required_product_type='Mobile AMR'.")
            lines.append("  (3) Everything else (pallets, IBCs, forks required, racking, heavy load) → required_product_type='Forklift AGV'.")
            lines.append("      Counterbalanced, Reach Truck, AND VNA are all 'Forklift AGV'.")
            lines.append("")
    _valid_values = ", ".join(repr(n) for n in names) if names else "null"
    lines += [
        "Fields:",
        f"- required_product_type: MANDATORY — exactly one of: {_valid_values}.",
        "",
        "DOCUMENT:",
        "{text}",
        "",
        "JSON:",
        '{"required_product_type":null}',
    ]
    content = "\n".join(lines)
    if dom_variant_guides:
        expected_names = list(dom_variant_guides.keys())
        assert all(n in content for n in expected_names), \
            f"classification template missing variant names: {expected_names}"
    elif leaf_scopes:
        expected_names = [n["canonical_name"] for n in leaf_scopes]
        assert all(n in content for n in expected_names), \
            f"classification template missing canonical names: {expected_names}"
    return content


# Fields determined in Pass 4a — excluded from Pass 4b templates
_4A_FIELDS = {"required_product_type"}


_OPERATOR_DIRECTION = {
    "KO_IF_LT": "CONSERVATIVE EXTRACTION: if multiple values are stated, extract the MAXIMUM — the supplier must meet or exceed this threshold.",
    "KO_IF_GT": "CONSERVATIVE EXTRACTION: if multiple values are stated, extract the MINIMUM — the supplier must not exceed this constraint.",
}
_NUMERIC_DTYPES = {"Float", "Integer"}


def build_extraction_template(vehicle_types: dict, extraction_schema: list,
                               scope_filter: str = None, field_levels: dict = None,
                               shared_scope: str = "", resolution=None,
                               scope_nodes: dict = None) -> str:
    """Build LLM extraction template.

    scope_filter=None  → full combined template (backward compat / JSON-retry fallback).
    scope_filter=<scope_id> → Pass 4b type-specific template; includes shared scope + that scope only,
                              excludes fields already determined in Pass 4a.
                              Placeholder {vehicle_type} filled by app.py at runtime.
    """
    if scope_filter:
        # Pass 4b: fields for this vehicle type (shared + type-specific + global if resolution given)
        _filter_set = set(resolution) if resolution else {shared_scope, scope_filter}
        schema_to_use = [
            f for f in extraction_schema
            if f.get("scope") in _filter_set
            and f["key"] not in _4A_FIELDS
        ]
        lines = [
            "The vehicle type for this tender has been determined: {vehicle_type}.",
            "Extract the technical requirements listed below from the tender document.",
            "Values may appear in running text OR in tables — extract from both.",
            "Output ONLY the JSON object shown at the end — nothing else.",
            "",
        ]
    else:
        # Full combined template (backward compat)
        schema_to_use = extraction_schema
        lines = ["Extract AGV/AMR technical requirements from this tender. Values may appear in running text OR in tables — extract from both.",
                 "The document may be written in any language (German, English, French, etc.). Extract fields based on their semantic meaning, not on specific keywords or section headings.",
                 ""]
        if scope_nodes:
            leaf_scopes = [n for n in scope_nodes.values() if n.get("canonical_name") and n.get("classification_guide")]
            if leaf_scopes:
                lines += ["Vehicle type classification guide (for required_product_type):"]
                for scope in leaf_scopes:
                    lines.extend(_render_leaf_scope_lines(scope))
                lines.append("  Key: PAYLOAD AND LOAD TYPE determine the vehicle — not the environment alone.")
                lines.append("  Filling lines + pallet transport = Forklift AGV. Filling lines + light totes/boxes = Mobile AMR.")
                lines.append("")
                lines.append("  THINK STEP BY STEP when classifying required_product_type:")
                lines.append("  IMPORTANT: Only three values are valid for required_product_type: 'Forklift AGV', 'Tugger AGV', 'Mobile AMR'.")
                lines.append("  Sub-variants (Counterbalanced, Reach Truck, VNA) are NOT valid outputs — they are internal properties, not types.")
                lines.append("  (1) Does the doc mention towing / tugger / milk run / trailer train / Routenzug? → required_product_type='Tugger AGV'.")
                lines.append("  (2) Is this light load (<1000 kg) with flexible SLAM navigation and no standard floor-pallet pickup? → required_product_type='Mobile AMR'.")
                lines.append("  (3) Everything else (pallets, IBCs, drums, racking, forks required, heavy load) → required_product_type='Forklift AGV'.")
                lines.append("      This includes Counterbalanced, Reach Truck, AND VNA applications — they are all 'Forklift AGV'.")
                lines.append("      If VNA / very narrow aisle / aisle<2m is detected → required_product_type='Forklift AGV' AND required_vna_capable=true.")
                lines.append("  Do NOT output 'Counterbalanced', 'Reach Truck', 'VNA', or any other sub-variant as the vehicle type.")
                lines.append("")
        else:
            guide = vehicle_types.get("llm_guide", [])
            if guide:
                lines += ["Vehicle type classification guide (for required_product_type):"]
                seen_names = set()
                for vt in guide:
                    if vt["name"] in seen_names: continue
                    seen_names.add(vt["name"])
                    line = f'  * "{vt["name"]}" → {vt["description"]}'
                    if vt.get("key_indicators"):
                        line += f'. Signals: {vt["key_indicators"]}'
                    lines.append(line)
                lines.append("  Key: PAYLOAD AND LOAD TYPE determine the vehicle — not the environment alone.")
                lines.append("  Filling lines + pallet transport = Forklift AGV. Filling lines + light totes/boxes = Mobile AMR.")
                lines.append("")
                lines.append("  THINK STEP BY STEP when classifying required_product_type:")
                lines.append("  IMPORTANT: Only three values are valid for required_product_type: 'Forklift AGV', 'Tugger AGV', 'Mobile AMR'.")
                lines.append("  Sub-variants (Counterbalanced, Reach Truck, VNA) are NOT valid outputs — they are internal properties, not types.")
                lines.append("  (1) Does the doc mention towing / tugger / milk run / trailer train / Routenzug? → required_product_type='Tugger AGV'.")
                lines.append("  (2) Is this light load (<1000 kg) with flexible SLAM navigation and no standard floor-pallet pickup? → required_product_type='Mobile AMR'.")
                lines.append("  (3) Everything else (pallets, IBCs, drums, racking, forks required, heavy load) → required_product_type='Forklift AGV'.")
                lines.append("      This includes Counterbalanced, Reach Truck, AND VNA applications — they are all 'Forklift AGV'.")
                lines.append("      If VNA / very narrow aisle / aisle<2m is detected → required_product_type='Forklift AGV' AND required_vna_capable=true.")
                lines.append("  Do NOT output 'Counterbalanced', 'Reach Truck', 'VNA', or any other sub-variant as the vehicle type.")
                lines.append("")

    lines.append("Field definitions:")
    _source_instrumented = set()  # numeric KO fields that get a _source companion
    for field in schema_to_use:
        if not field["hint"]:
            continue
        mand = " MANDATORY —" if field["mandatory"] else ""
        hint = field["hint"]
        needs_source = False
        if field_levels:
            fl  = field_levels.get(field["db_field"], {})
            op  = fl.get("operator", "")
            dt  = fl.get("data_type", "")
            if op in _OPERATOR_DIRECTION and dt in _NUMERIC_DTYPES:
                hint = hint + " " + _OPERATOR_DIRECTION[op]
                needs_source = True
        lines.append(f'- {field["key"]}:{mand} {hint}')
        if field.get("variant_scope") and scope_nodes:
            _variant_guides = scope_nodes.get(field["variant_scope"], {}).get("variant_guides", {})
            for variant, guide in _variant_guides.items():
                lines.append(f'  * "{variant}" → {guide}')
        if needs_source:
            _source_instrumented.add(field["key"])
            lines.append(
                f'- {field["key"]}_source: The EXACT verbatim sentence or phrase from the document '
                f'that contains the value above. null if the value was not found explicitly in the '
                f'document — in which case {field["key"]} MUST also be null.'
            )

    lines += ["",
              "Use null only when a field genuinely cannot be determined from the document.",
              "",
              "DOCUMENT:",
              "{text}",
              "",
              "JSON:"]

    seen_keys = set()
    json_fields = []
    for field in schema_to_use:
        if field["key"] not in seen_keys:
            seen_keys.add(field["key"])
            json_fields.append(f'"{field["key"]}":null')
            if field["key"] in _source_instrumented:
                json_fields.append(f'"{field["key"]}_source":null')
    lines.append("{" + ",".join(json_fields) + "}")

    return "\n".join(lines)


def build_retry_template(extraction_schema: list) -> str:
    seen_keys = set()
    json_fields = []
    for field in extraction_schema:
        if field["key"] not in seen_keys:
            seen_keys.add(field["key"])
            json_fields.append(f'"{field["key"]}":null')
    json_schema = "{" + ",".join(json_fields) + "}"

    return "\n".join([
        "IMPORTANT: Output ONLY the JSON object below, nothing else. No prose, no explanation.",
        "",
        "Extract AGV/AMR requirements from this tender. Use null for values not explicitly stated.",
        "",
        "Key interpretation rules:",
        "- If a handling point table lists \"Conveyor belt picking/delivery\" as the handling method, set required_conveyor_integration to \"yes\".",
        "- Tugger AGVs tow trailer trains and CANNOT interface with conveyor belts without manual reloading — if conveyors are present, required_product_type should NOT be Tugger.",
        "- If most stations use floor delivery but at least one uses conveyor belt, the system needs BOTH floor transport AND conveyor integration.",
        "",
        "DOCUMENT:",
        "{text}",
        "",
        "Output ONLY this JSON:",
        json_schema,
    ])


def build_nace_template(nace: dict) -> str:
    return "\n".join([
        'A tender is for: "{tender_category}"',
        "",
        "SCOPE RULE: Evaluate ONLY what is being PROCURED (the tendered service/product).",
        "The buyer's industry is IRRELEVANT — an AGV system tender from a beverage company",
        "is IN SCOPE just as much as one from a logistics company.",
        "",
        f'IN SCOPE: {nace["scope_in"]}.',
        "",
        f'OUT OF SCOPE: {nace["scope_out"]}.',
        "",
        "If OUT OF SCOPE, output the not-in-scope JSON.",
        "Otherwise, pick the single best matching NACE code from this list:",
        "{category_list}",
        "",
        "Output ONLY one of these two JSON options — use exactly these field names:",
        "",
        'If in scope:',
        '{"nace_tender":"[code]","nace_tender_name":"[name from list]","nace_buyer":null,"priority":"TOP|GUT|NIEDRIG","confidence":"HOCH|MITTEL|NIEDRIG","in_scope":true}',
        "",
        "If out of scope:",
        '{"nace_tender":null,"nace_tender_name":null,"nace_buyer":null,"priority":"UNBEKANNT","confidence":"NIEDRIG","in_scope":false}',
    ])


def build_basic_template() -> str:
    return "\n".join([
        "Extract these facts from the tender document. Output ONLY valid JSON.",
        "Use JSON null (not the string \"null\") if a value is not found.",
        "Dates: extract in ANY format found; normalize output to DD.MM.YYYY.",
        "",
        "Rules:",
        "- buyer_industry: infer from the buyer's name/description if not explicitly stated",
        "  (e.g. \"Technische Universität\" → \"Higher Education\", a hospital → \"Healthcare\").",
        "  Never leave null if the buyer name gives a clear hint.",
        "- tender_category: describe what is being procured in 3-6 words",
        "  (e.g. \"Website Redesign\", \"Warehouse Automation System\", \"Security Consulting\").",
        "  Never leave null if the project name is present.",
        "",
        "DOCUMENT:",
        "{text}",
        "",
        "JSON:",
        '{"buyer":null,"project_name":null,"project_location":null,"tender_date":null,"deadline":null,"contact_name":null,"contact_email":null,"contact_phone":null,"buyer_industry":null,"tender_category":null,"summary":null,"missing_info":[]}',
    ])


def build_domain_detection_template(scope_nodes: dict, scopes_raw: list) -> str:
    """Pass 2b: classify which haystacked domain this tender belongs to.

    scope_nodes carries per-leaf info (keywords, classification_guide) but is empty
    for domain-level nodes. Use scopes_raw (sid, parent, tab_name tuples) to identify
    domain-level scopes (direct children of root '*' with a tab_name).
    """
    domain_scope_ids = [sid for sid, p, t in scopes_raw if p == "*" and t]

    # Aggregate keywords from leaf children — domain-level nodes have no own keywords.
    _parent_of = {sid: p for sid, p, _ in scopes_raw}
    domain_kws: dict = {}
    for domain_sid in domain_scope_ids:
        child_kws = [
            kw
            for cnode_sid, cnode in scope_nodes.items()
            if _parent_of.get(cnode_sid) == domain_sid
            for kw in cnode.get("scope_keywords", [])
        ]
        domain_kws[domain_sid] = list(dict.fromkeys(child_kws))

    lines = [
        "Classify which procurement domain this tender belongs to.",
        "Base your answer ONLY on the tender_category and summary provided.",
        "Output ONLY valid JSON — nothing else.",
        "",
    ]
    if domain_scope_ids:
        lines.append("Available domains:")
        for sid in domain_scope_ids:
            kws_sample = ", ".join(domain_kws.get(sid, [])[:8])
            lines.append(f'  * "{sid}". Keywords: {kws_sample}')
        valid_ids = [repr(sid) for sid in domain_scope_ids]
        lines.append(f"")
        lines.append(f"Set detected_domain to one of: {', '.join(valid_ids)}, or null if none match.")
    lines += [
        "",
        "TENDER_CATEGORY: {tender_category}",
        "SUMMARY: {summary}",
        "",
        "JSON:",
        '{"detected_domain": null}',
    ]
    return "\n".join(lines)


# ── Validators ────────────────────────────────────────────────────────────────

def _scope_resolution_chain(scope_id: str, parent_of: dict) -> list:
    """Walk the scope tree from root to leaf; return all scope_ids in order.

    Used by both _write_scope_registry (to store resolution_order) and
    generate() (to select fields for VT-specific extraction templates) —
    single implementation eliminates the SA-06 drift risk.
    """
    chain, cur = [], scope_id
    while cur:
        chain.insert(0, cur)
        cur = parent_of.get(cur)
    return chain


def validate_unit_suffix_drift(fields: dict) -> list:
    """OI-55: warn when a field_name unit suffix disagrees with the AP0 unit column.

    Example: lifting_height_mm has suffix '_mm' but AP0 unit='m' after the
    unit-conversion refactor. This is a documentation smell, not a runtime error.
    The check is deliberately a warning (not an assert) so generation is never blocked.
    Suffix aliases (e.g. 'h' == 'hours') are normalised before comparison.
    """
    _SUFFIX_UNIT = {"_mm": "mm", "_kg": "kg", "_h": "h", "_min": "min", "_pct": "%"}
    _ALIASES = {"hours": "h", "meter": "m", "metres": "m", "meters": "m"}
    warnings = []
    for uuid, f in fields.items():
        fn = f.get("field_name", "")
        raw_unit = (f.get("unit") or "").strip()
        unit = _ALIASES.get(raw_unit.lower(), raw_unit)
        for suffix, expected in _SUFFIX_UNIT.items():
            if fn.endswith(suffix) and unit and unit != expected:
                warnings.append(
                    f"Unit-suffix drift: '{fn}' ends in '{suffix}' "
                    f"but AP0 unit='{raw_unit}' — rename field_name or update AP0 unit column"
                )
                break
    return warnings


def validate_vs_sqlite(field_levels: dict, db_path: Path) -> list:
    if not db_path.exists():
        return ["SQLite DB not found — schema validation skipped"]
    conn = sqlite3.connect(str(db_path))
    db_cols = set()
    for table in ("companies", "products", "base_model_extensions"):
        db_cols.update(r[1] for r in conn.execute(f"PRAGMA table_info({table})").fetchall())
    conn.close()
    db_cols -= SQLITE_SKIP
    warnings = []
    for f in sorted(set(field_levels) - db_cols):
        warnings.append(f"AP0 defines '{f}' ({field_levels[f]['level']}) but missing from SQLite")
    for f in sorted(db_cols - set(field_levels)):
        warnings.append(f"SQLite has '{f}' with no AP0 classification")
    return warnings




# ── Fields JSON + field_spec.py ───────────────────────────────────────────────

def _check_null_rule_marker(fname, hint) -> None:
    """Fail loud if `hint` contains the substring 'NULL RULE' in a spelling that
    does not match the canonical marker pattern (src.prompt_markers.NULL_RULE_PATTERN).
    An unrecognised marker variant would silently leak null-bias prose into Pass 4c's
    prompt undiluted — see OI-109 D2."""
    if hint and "NULL RULE" in hint and not NULL_RULE_PATTERN.search(hint):
        sys.exit(
            f"[FEHLER] emit_fields_json: Feld '{fname}' enthält 'NULL RULE' in einer "
            "nicht erkannten Marker-Form (erwartet z.B. 'NULL RULE:' oder "
            "'NULL RULE — READ FIRST:'). Siehe src/prompt_markers.NULL_RULE_PATTERN."
        )


def emit_fields_json(wb, data_sheets, plausibility_raw=None, tab_scope_map: dict = None, scope_nodes: dict = None) -> None:
    """Read all sheets from data_sheets and write config/fields.json + src/field_spec.py.

    fields.json is keyed by UUID. field_spec.py is a generated Python module with
    FieldSpec dataclass and helper accessors.
    """
    if plausibility_raw is None:
        plausibility_raw = {}
    if tab_scope_map is None:
        tab_scope_map = {}

    def _nullable_float(v):
        if v is None or v == "":
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    def _nullable_int(v):
        if v is None or v == "":
            return None
        try:
            return int(v)
        except (TypeError, ValueError):
            return None

    def _snake(s: str) -> str:
        return s.strip().lower().replace(" ", "_").replace("-", "_")

    fields_by_uuid: dict = {}

    for sheet in data_sheets:
        if sheet not in wb.sheetnames:
            print(f"  [WARN] Sheet missing for emit_fields_json: {sheet}")
            continue
        rows = _rows(wb[sheet])
        hi, cols = _find_header(rows)
        if hi is None:
            continue

        # Normalise column headers to snake_case for robust lookup
        cols_snake = {_snake(k): v for k, v in cols.items()}

        col_uuid          = cols.get("UUID")
        col_fname         = cols.get("Field Name", 0)
        col_entity        = cols.get("Entity")
        col_level         = cols.get("Level")
        col_op            = cols.get("Matching Operator")
        col_dtype         = cols.get("Data Type")
        col_unit          = cols.get("Unit")
        col_allowed       = cols.get("Allowed Values")
        col_sfunc         = cols.get("Score Function")
        col_ta            = cols.get("Score Threshold A")
        col_tb            = cols.get("Score Threshold B")
        col_weight        = cols.get("Scoring Weight")
        col_hint          = cols.get("LLM Hint")
        col_display       = cols.get("Display Mode")
        col_client        = cols.get("UI Hint")
        col_display_label = cols.get("Display Label")
        col_value_if_null = cols_snake.get("value_if_null")

        for row in rows[hi + 1:]:
            fname = row[col_fname] if col_fname < len(row) else None
            if not fname or str(fname).startswith("──"):
                continue

            uuid_val = row[col_uuid] if col_uuid is not None and col_uuid < len(row) else None
            if not uuid_val:
                sys.exit(
                    f"[FEHLER] emit_fields_json: Feld '{fname}' in Sheet '{sheet}' hat keine UUID. "
                    "Alle Felder müssen eine UUID haben."
                )
            uuid_str = str(uuid_val).strip()

            if uuid_str in fields_by_uuid:
                existing = fields_by_uuid[uuid_str]
                sys.exit(
                    f"[FEHLER] emit_fields_json: UUID-Kollision! '{uuid_str}' erscheint in "
                    f"Sheet '{sheet}' (field='{fname}') UND in Sheet '{existing.get('scope', '')}' "
                    f"(field='{existing['field_name']}'). UUIDs müssen eindeutig sein."
                )

            tkey = "required_" + str(fname).strip()
            entity = str(row[col_entity]).strip() if col_entity is not None and col_entity < len(row) and row[col_entity] else ""
            raw_level = row[col_level] if col_level is not None and col_level < len(row) else None
            level = LEVEL_MAP.get(str(raw_level).strip()) if raw_level else None
            op = str(row[col_op]).strip() if col_op is not None and col_op < len(row) and row[col_op] else None
            dtype = str(row[col_dtype]).strip() if col_dtype is not None and col_dtype < len(row) and row[col_dtype] else ""
            unit = str(row[col_unit]).strip() if col_unit is not None and col_unit < len(row) and row[col_unit] else None
            sfunc = str(row[col_sfunc]).strip() if col_sfunc is not None and col_sfunc < len(row) and row[col_sfunc] else None
            ta = _nullable_float(row[col_ta] if col_ta is not None and col_ta < len(row) else None)
            tb = _nullable_float(row[col_tb] if col_tb is not None and col_tb < len(row) else None)
            weight = _nullable_int(row[col_weight] if col_weight is not None and col_weight < len(row) else None)
            hint = str(row[col_hint]).strip() if col_hint is not None and col_hint < len(row) and row[col_hint] else None
            _check_null_rule_marker(fname, hint)
            display = str(row[col_display]).strip() if col_display is not None and col_display < len(row) and row[col_display] else None
            client_exp = str(row[col_client]).strip() if col_client is not None and col_client < len(row) and row[col_client] else None
            display_label = str(row[col_display_label]).strip() if col_display_label is not None and col_display_label < len(row) and row[col_display_label] else None

            # Parse allowed_values only for Dropdown/Multi-Select — numeric fields
            # store unit strings in this column which must not be treated as enums.
            allowed = None
            if dtype in ("Dropdown", "Multi-Select") and col_allowed is not None and col_allowed < len(row) and row[col_allowed]:
                raw_av = str(row[col_allowed]).strip()
                if isinstance(raw_av, str) and raw_av.startswith("@SCOPE_VARIANTS:"):
                    scope_ref = raw_av.split(":", 1)[1].strip()  # e.g. "FoodBev:Refrigeration"
                    _sn = scope_nodes or {}
                    av_list = (_sn.get(scope_ref) or {}).get("scope_variants", [])
                    if not av_list:
                        print(f"  [WARN] @SCOPE_VARIANTS:{scope_ref} — scope_variants empty or scope not found")
                elif raw_av == "@SCOPE_CANONICAL_NAMES":
                    # Expand sentinel: sorted canonical_names of all scope nodes
                    _sn = scope_nodes or {}
                    av_list = sorted(
                        n["canonical_name"] for n in _sn.values() if n.get("canonical_name")
                    )
                else:
                    av_list = [v.strip() for v in raw_av.split("|") if v.strip() and v.strip() not in ("…", "")]
                if av_list:
                    allowed = av_list

            # Parse Value if Null — typed coercion with validation
            raw_vin = str(row[col_value_if_null]).strip() if col_value_if_null is not None and col_value_if_null < len(row) and row[col_value_if_null] else ""
            value_if_null = None
            if raw_vin:
                fname_str = str(fname).strip()
                if dtype == "Boolean":
                    if raw_vin.lower() not in ("true", "false"):
                        sys.exit(f"ERROR: Value if Null for '{fname_str}' is '{raw_vin}' but data_type is Boolean — must be 'true' or 'false'")
                    value_if_null = raw_vin.lower() == "true"
                elif dtype in ("Float", "Integer"):
                    try:
                        value_if_null = float(raw_vin) if dtype == "Float" else int(float(raw_vin))
                    except ValueError:
                        sys.exit(f"ERROR: Value if Null for '{fname_str}' is '{raw_vin}' but data_type is {dtype} — must be numeric")
                    plaus = plausibility_raw.get("required_" + fname_str)
                    if plaus:
                        lo, hi_val = plaus.get("min"), plaus.get("max")
                        if lo is not None and value_if_null < lo:
                            sys.exit(f"ERROR: Value if Null for '{fname_str}' = {value_if_null} is below plausibility min {lo}")
                        if hi_val is not None and value_if_null > hi_val:
                            sys.exit(f"ERROR: Value if Null for '{fname_str}' = {value_if_null} is above plausibility max {hi_val}")
                elif dtype in ("String", "Single-select", "Dropdown"):
                    if allowed and raw_vin not in allowed:
                        sys.exit(f"ERROR: Value if Null for '{fname_str}' is '{raw_vin}' — not in allowed_values {allowed}")
                    value_if_null = raw_vin
                elif dtype == "Multi-Select":
                    parts = [p.strip() for p in raw_vin.split("|")]
                    if allowed:
                        bad = [p for p in parts if p not in allowed]
                        if bad:
                            sys.exit(f"ERROR: Value if Null for '{fname_str}': {bad} not in allowed_values {allowed}")
                    value_if_null = parts
                else:
                    value_if_null = raw_vin

            fields_by_uuid[uuid_str] = {
                "uuid":             uuid_str,
                "field_name":       str(fname).strip(),
                "tender_key":       tkey,
                "entity":           entity,
                "scope":            tab_scope_map.get(sheet, sheet),
                "level":            level,
                "operator":         op,
                "data_type":        dtype,
                "unit":             unit,
                "allowed_values":   allowed,
                "score_function":   sfunc,
                "threshold_a":      ta,
                "threshold_b":      tb,
                "scoring_weight":    weight,
                "hint":             hint,
                "user_description": client_exp,
                "display_mode":     display,
                "display_label":    display_label,
                "value_if_null":    value_if_null,
            }

    print(f"  Fields JSON: {len(fields_by_uuid)} entries across {len(data_sheets)} sheets")

    # SA-22: CONTEXT fields must not have an operator — fail generation if violated.
    context_with_op = [
        f"{v['field_name']} (op={v['operator']!r})"
        for v in fields_by_uuid.values()
        if v.get("level") == "CONTEXT" and v.get("operator")
    ]
    if context_with_op:
        sys.exit(
            f"[FEHLER] AP0 hygiene: CONTEXT-level fields must not have an operator. "
            f"Fix in AP0 xlsx:\n  " + "\n  ".join(context_with_op)
        )

    # SA-25: Warn about SCORING fields that carry no weight and no score_function — they are inert.
    inert_scoring = [
        v["field_name"]
        for v in fields_by_uuid.values()
        if v.get("level") == "SCORING" and not v.get("scoring_weight") and not v.get("score_function")
    ]
    if inert_scoring:
        print(f"  [SA-25 WARN] {len(inert_scoring)} SCORING-level fields carry no weight/score_function "
              f"— inert in matching. Assign weights in AP0 or demote to CONTEXT.")

    # Collision Guard: same field_name in multiple entries → (data_type, unit) must be identical.
    # Exception: Global-scope fields (scope == "*") are allowed to appear in multiple domains.
    seen: dict[str, tuple] = {}
    for entry in fields_by_uuid.values():
        name = entry.get("field_name", "")
        scope = entry.get("scope", "")
        if scope == "*":
            continue
        key = (entry.get("data_type"), entry.get("unit", ""))
        if name in seen and seen[name] != key:
            raise SystemExit(
                f"Field name collision: '{name}' has conflicting (data_type, unit): "
                f"{seen[name]} vs {key}"
            )
        seen[name] = key

    # Write config/fields.json
    (CONFIG_DIR / "fields.json").write_text(
        json.dumps(fields_by_uuid, indent=2, ensure_ascii=False) + "\n"
    )

    # Write src/field_spec.py as a string template
    field_spec_src = '''\
# src/field_spec.py  — GENERATED by scripts/generate_all.py — do not edit manually
from __future__ import annotations
from dataclasses import dataclass
from typing import Optional
import json
from pathlib import Path

ROOT = Path(__file__).parent.parent
_FIELDS_JSON = ROOT / "config" / "fields.json"


@dataclass
class FieldSpec:
    uuid: str
    field_name: str
    tender_key: Optional[str]
    entity: str
    scope: str
    level: Optional[str]
    operator: Optional[str]
    data_type: str
    unit: Optional[str]
    allowed_values: Optional[list]
    score_function: Optional[str]
    threshold_a: Optional[float]
    threshold_b: Optional[float]
    scoring_weight: Optional[int]
    hint: Optional[str]
    user_description: Optional[str]
    display_mode: Optional[str]
    display_label: Optional[str] = None
    value_if_null: object = None


def load_fields() -> dict[str, FieldSpec]:
    """Returns all fields keyed by UUID. Asserts UUID uniqueness."""
    data = json.loads(_FIELDS_JSON.read_text())
    result = {k: FieldSpec(**v) for k, v in data.items()}
    assert len(result) == len(data), "Duplicate UUIDs in fields.json"
    return result


def fields_by_tender_key() -> dict[str, list[FieldSpec]]:
    """Returns fields grouped by tender_key. One key can map to multiple VT-scoped fields.
    Only includes fields where tender_key is not None."""
    result: dict[str, list[FieldSpec]] = {}
    for f in load_fields().values():
        if f.tender_key:
            result.setdefault(f.tender_key, []).append(f)
    return result


def fields_by_field_name() -> dict[str, list[FieldSpec]]:
    """Returns fields grouped by field_name. One name can appear in multiple VT scopes."""
    result: dict[str, list[FieldSpec]] = {}
    for f in load_fields().values():
        result.setdefault(f.field_name, []).append(f)
    return result


'''
    (ROOT / "src" / "field_spec.py").write_text(field_spec_src, encoding="utf-8")


def _write_scope_registry(config_dir: Path, scopes_raw: list, tab_scope_map: dict, vt_map: dict,
                           scope_nodes: dict, dry_run: bool = False) -> dict:
    """Write config/scope_registry.json from scope registry data. Returns legacy_map."""
    parent_of = {sid: parent for sid, parent, _ in scopes_raw}
    tab_of    = {sid: tab for sid, _, tab in scopes_raw}

    def resolution_chain(scope_id: str) -> list:
        return _scope_resolution_chain(scope_id, parent_of)

    parent_ids    = {parent for _, parent, _ in scopes_raw if parent}
    leaf_scope_ids = [sid for sid, _, _ in scopes_raw if sid not in parent_ids]

    scopes = {}
    for sid, _, _ in scopes_raw:
        node = {
            "scope_id": sid,
            "parent": parent_of.get(sid) or None,
            "tab_name": tab_of.get(sid, ""),
        }
        sn = scope_nodes.get(sid, {})
        for field in ("canonical_name", "classification_guide", "scope_variants", "scope_keywords",
                      "vna_applicable", "scoring_bucket", "vna_hint"):
            v = sn.get(field)
            if v is not None:
                node[field] = v
        scopes[sid] = node

    # Derive variant_map: variant.lower() → canonical_name (from scope_variants × canonical_name)
    variant_map: dict = {}
    for node in scopes.values():
        cn = node.get("canonical_name")
        if not cn:
            continue
        for variant in node.get("scope_variants", []):
            variant_map[variant.strip().lower()] = cn

    # Derive agv_detection_keywords: union of scope_keywords of all nodes with canonical_name
    agv_detection_keywords: list = list(dict.fromkeys(
        kw
        for node in scopes.values()
        if node.get("canonical_name")
        for kw in node.get("scope_keywords", [])
    ))

    # Derive domain_keywords: domain-node own keywords + union of all child scope_keywords
    domain_keywords: dict = {}
    for sid, node in scopes.items():
        if node.get("parent") == "*" and node.get("tab_name"):
            own_kws = node.get("scope_keywords", [])
            child_kws = [
                kw
                for csid, cnode in scopes.items()
                if cnode.get("parent") == sid
                for kw in cnode.get("scope_keywords", [])
            ]
            domain_keywords[sid] = list(dict.fromkeys(own_kws + child_kws))

    # Build legacy_map: canonical_name → scope_id, plus scope_variants → scope_id
    # (scope_variants lets IK sub-categories like "Cold Store" route to FoodBev:Refrigeration)
    legacy_map: dict = {}
    for node in scopes.values():
        cn  = node.get("canonical_name")
        sid = node["scope_id"]
        if cn:
            legacy_map[cn] = sid
        for _v in node.get("scope_variants", []):
            v = _v.strip()
            if v and v not in legacy_map:
                legacy_map[v] = sid

    # Parent-conformance assertion
    _declared = set(scopes.keys())
    _roots = [sid for sid, node in scopes.items() if node.get("parent") is None]
    assert len(_roots) == 1, f"Expected exactly one root scope, got: {_roots}"
    for sid, node in scopes.items():
        p = node.get("parent")
        if p is not None:
            assert p in _declared, f"scope {sid!r} declares parent {p!r} which is not a declared scope_id"
            prefix = sid.rsplit(":", 1)[0] if ":" in sid else None
            expected_parent = prefix if prefix in _declared else "*"
            assert p == expected_parent, \
                f"scope {sid!r}: declared parent {p!r} != derived parent {expected_parent!r}"

    # Variant superset assertion
    _PRE_STEP7_VT_MAP_KEYS = frozenset([
        "vna", "very narrow aisle", "reach truck", "counterbalanced", "forklift",
        "forklift agv", "agv", "tugger", "tugger agv", "mobile amr", "amr",
        "underride amr", "underride", "autonomous mobile robot"
    ])
    assert _PRE_STEP7_VT_MAP_KEYS <= set(variant_map.keys()), \
        f"variant_map missing pre-Step-7 vt_map entries: {_PRE_STEP7_VT_MAP_KEYS - set(variant_map.keys())}"

    # Keyword superset assertion
    _PRE_STEP7_DETECT_KWS = frozenset([
        "vna", "very narrow aisle", "reach truck", "schmalgangstapler", "counterbalanced",
        "forklift", "stapler", "gabelstapler", "forklift agv", "agv", "tugger", "schlepper",
        "routenzug", "milk run", "towing", "anhaenger", "tugger agv", "amr", "mobile robot",
        "fahrroboter", "autonomer", "goods-to-person", "goods to person", "underride amr",
        "unterfahrfahrzeug", "underride", "autonomous mobile robot"
    ])
    assert _PRE_STEP7_DETECT_KWS <= set(agv_detection_keywords), \
        f"agv_detection_keywords missing pre-Step-7 entries: {_PRE_STEP7_DETECT_KWS - set(agv_detection_keywords)}"

    resolution_order = {
        leaf_sid: resolution_chain(leaf_sid)
        for leaf_sid in leaf_scope_ids
    }

    result = {
        "scopes": scopes,
        "resolution_order": resolution_order,
        "legacy_map": legacy_map,
        "variant_map": variant_map,
        "agv_detection_keywords": agv_detection_keywords,
        "domain_keywords": domain_keywords,
    }

    if not dry_run:
        (config_dir / "scope_registry.json").write_text(
            json.dumps(result, indent=2, ensure_ascii=False) + "\n"
        )
    print(f"  scope_registry.json: {len(scopes)} scopes, {len(leaf_scope_ids)} leaf scopes, {len(legacy_map)} legacy mappings")
    return legacy_map


# ── Main ──────────────────────────────────────────────────────────────────────

def generate(xlsx_path: Path, db_path: Path, dry_run: bool = False,
             platform_path: Path = DEFAULT_PLATFORM) -> int:
    print(f"Reading AP0 xlsx:      {xlsx_path.name}")
    print(f"Reading platform config: {platform_path.name if platform_path.exists() else '(not found)'}")
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)

    data_sheets, shared_tab, leaf_tabs, tab_scope_map, scopes_raw, scope_nodes = read_scope_registry(wb)

    field_levels         = read_field_levels(wb, data_sheets)
    vehicle_types        = read_vehicle_types(wb)
    field_text_fallbacks = read_field_text_fallbacks(wb)
    vehicle_types["field_text_fallbacks"] = field_text_fallbacks
    scoring_weights      = read_scoring_weights(wb, data_sheets)
    extraction_schema = read_extraction_schema(wb, data_sheets, tab_scope_map=tab_scope_map)
    sqlite_schema     = read_sqlite_schema(wb, data_sheets)
    plausibility_raw  = read_plausibility(wb, data_sheets)

    wb_platform       = openpyxl.load_workbook(str(platform_path), read_only=True, data_only=True) \
                        if platform_path.exists() else None
    unit_conversions  = read_unit_conversions(wb_platform) if wb_platform else {}
    plausibility      = build_plausibility_config(plausibility_raw, unit_conversions)

    # Assert all numeric KO fields (Float/Integer, KO_IF_LT/KO_IF_GT) have plausibility
    # ranges defined in the AP0 xlsx. Missing entries cause silent validation gaps.
    _numeric_ko_keys = {
        meta["tender_key"]
        for meta in field_levels.values()
        if meta.get("operator") in ("KO_IF_LT", "KO_IF_GT")
        and meta.get("data_type") in ("Float", "Integer")
    }
    _missing = _numeric_ko_keys - set(plausibility_raw)
    assert not _missing, (
        f"[FEHLER] Numeric KO-Felder ohne Plausibility-Daten in AP0 xlsx: {_missing}. "
        "Plausibility Min / Plausibility Max / Unit in der AP0-Tabelle ergänzen."
    )

    platform          = read_platform(wb_platform, platform_path)
    nace              = platform  # nace data is inside platform dict

    # OI-98: contact-fallback pass field set, derived from the "Contact Fallback"
    # column of the Basic Extraction Schema tab (platform_config.xlsx). "trigger"
    # fields decide whether the fallback pass runs; "target" fields (plus all
    # trigger fields, which are always also targets) are the fields it may fill in.
    _contact_trigger_keys = [
        e["key"] for e in platform["basic_schema"] if e.get("contact_fallback") == "trigger"
    ]
    _contact_target_keys = _contact_trigger_keys + [
        e["key"] for e in platform["basic_schema"] if e.get("contact_fallback") == "target"
    ]
    if not _contact_trigger_keys:
        sys.exit(
            "[FEHLER] Basic Extraction Schema: kein Feld mit Contact Fallback='trigger' — "
            "der Kontakt-Fallback-Pass würde nie ausgelöst. contact_name/contact_email/"
            "contact_phone (oder Äquivalent) in platform_config.xlsx markieren."
        )

    # Additional runtime fields derived from AP0.
    # scoring_bucket_map, vna_applicable_types, agv_detection_keywords retired in Step 7:
    # these are now emitted to scope_registry.json via _write_scope_registry().

    # M-1: ensure Schmalgangstapler text override sets vna=True
    _schmal_regex = "(?i)schmalgangstapler"
    _existing = {o.get("regex") for o in vehicle_types.get("text_overrides", [])}
    if _schmal_regex not in _existing:
        vehicle_types.setdefault("text_overrides", []).append(
            {"regex": _schmal_regex, "canonical": "Forklift AGV", "vna": True}
        )

    print(f"  Fields: {len(field_levels)} ({sum(1 for v in field_levels.values() if v['level']=='KO')} KO, {sum(1 for v in field_levels.values() if v['level']=='COND_KO')} COND_KO)")
    print(f"  Vehicle types: {len(vehicle_types.get('vt_map', {}))} mappings, {len(vehicle_types.get('llm_guide',[]))} with LLM guide")
    print(f"  Scoring weights: {sum(len(v) for v in scoring_weights.values())} entries")
    print(f"  NACE codes (Prio 1): {len(platform['codes'])}")
    print(f"  Extraction fields: {len(extraction_schema)}")
    print(f"  SQLite schema: {sum(1 for s in sqlite_schema.values() if isinstance(s, str))} tables generated")
    print(f"  Plausibility ranges: {len(plausibility)} fields")

    # Build prompts
    shared_scope_id = tab_scope_map[shared_tab]
    vehicle_type_template = build_vehicle_type_template(vehicle_types, scope_nodes=scope_nodes)
    extraction_template   = build_extraction_template(vehicle_types, extraction_schema, field_levels=field_levels, shared_scope=shared_scope_id, scope_nodes=scope_nodes)
    retry_template        = build_retry_template(extraction_schema)

    # Step 4: compute scope→canonical mapping and resolution chains before building templates.
    # _write_scope_registry is called here (before dry_run check) so its return value is
    # available for the vt_prompt_map loop. File write is guarded by dry_run flag.
    legacy_map = _write_scope_registry(CONFIG_DIR, scopes_raw, tab_scope_map, vehicle_types.get("vt_map", {}), scope_nodes, dry_run=dry_run)
    # scope_id → canonical VT name (reverse of legacy_map)
    scope_to_canonical = {
        sid: n["canonical_name"]
        for sid, n in scope_nodes.items()
        if n.get("canonical_name")
    }
    # Build resolution chains for each leaf scope (walk up parent chain)
    _parent_of = {sid: p for sid, p, _ in scopes_raw}
    resolution_chains = {
        tab_scope_map[t]: _scope_resolution_chain(tab_scope_map[t], _parent_of)
        for t in leaf_tabs
        if t in tab_scope_map
    }

    # Build type-specific Pass 4b templates — derived from leaf_tabs, no hardcoded type names
    # vt_prompt_map local var still needed for file pruning; NOT written to vehicle_types.json (retired).
    vt_prompt_map: dict = {}
    type_templates: dict = {}
    for _sheet in leaf_tabs:
        _scope     = tab_scope_map.get(_sheet, _sheet)
        _canonical = scope_to_canonical.get(_scope, _sheet)
        _slug      = _sheet.lower().replace(" ", "_")
        _fname     = f"extraction_template_{_slug}.txt"
        vt_prompt_map[_canonical] = _fname
        type_templates[_sheet]    = build_extraction_template(
            vehicle_types, extraction_schema,
            scope_filter=_scope, field_levels=field_levels,
            shared_scope=shared_scope_id,
            resolution=resolution_chains.get(_scope),
            scope_nodes=scope_nodes,
        )
    vehicle_types["4a_fields"] = list(_4A_FIELDS)   # fields owned by Pass 4a, excluded from 4b
    # vt_prompt_map, vna_context_hint, scoring_bucket_map, vna_applicable_types,
    # agv_detection_keywords retired in Step 7 — now in scope_registry.json.
    nace_template             = build_nace_template(nace)
    basic_template            = build_basic_template()
    domain_detection_template = build_domain_detection_template(scope_nodes, scopes_raw)

    # Checksums & validation
    xlsx_md5 = hashlib.md5(xlsx_path.read_bytes()).hexdigest()
    warnings = validate_vs_sqlite(field_levels, db_path)
    drift_warnings = validate_unit_suffix_drift(
        json.loads((CONFIG_DIR / "fields.json").read_text()) if (CONFIG_DIR / "fields.json").exists() else {}
    )
    for w in drift_warnings:
        print(f"[OI-55 WARN] {w}")

    if dry_run:
        print("\n[DRY RUN] Would write:")
        print("  config/vehicle_types.json, nace_codes.json")
        print("  config/sqlite_schema.json, config/plausibility.json")
        print("  config/prompts/extraction_template.txt (and others)")
        print("  config/prompts/scope_classification_template.txt")
        print(f"\nExtraction template preview (first 400 chars):\n{extraction_template[:400]}")
        print(f"\nSQLite companies preview:\n{sqlite_schema['companies'][:400]}")
    else:
        PROMPTS_DIR.mkdir(parents=True, exist_ok=True)

        (CONFIG_DIR / "vehicle_types.json").write_text(
            json.dumps(vehicle_types, indent=2, ensure_ascii=False) + "\n")
        (CONFIG_DIR / "nace_codes.json").write_text(
            json.dumps(nace, indent=2, ensure_ascii=False) + "\n")
        (CONFIG_DIR / "sqlite_schema.json").write_text(
            json.dumps(sqlite_schema, indent=2, ensure_ascii=False) + "\n")
        (CONFIG_DIR / "plausibility.json").write_text(
            json.dumps(plausibility, indent=2, ensure_ascii=False) + "\n")

        (PROMPTS_DIR / "extraction_system.txt").write_text(
            "You are a technical procurement specialist. Extract technical requirements from tender documents. Output ONLY valid JSON. No markdown, no explanation.\n\n"
            "## Critical extraction rules\n"
            "8. CONSERVATIVE VALUE EXTRACTION (critical): When a document lists multiple values for the same parameter, always extract the most demanding value. "
            "For minimum-capability fields (payload, lift height, capacity, operating hours, fleet size, maximum ambient temperature): extract the MAXIMUM value found. "
            "For maximum-constraint fields (aisle width, minimum ambient temperature, minimum temperature setpoint): extract the MINIMUM value found. "
            "Never average or omit ambiguous values — always pick the worst case for the supplier.\n"
            "9. ANTI-HALLUCINATION (critical): Before outputting any non-null value you must be able to identify the exact sentence in the document that states it. "
            "Do NOT infer specifications from the product type, industry, or environment — stated context does NOT imply unstated technical values. "
            "Do NOT read numbers from dates, filenames, revision codes, version strings, or project metadata as specification values — "
            "'25th May 2022' is a date, NOT a temperature; 'v1.3' is a version, NOT a specification value. "
            "If a field's value is not directly stated in the document text, output null — never apply typical industry values."
        )
        # Per-domain classification templates (one per domain, domain-specific leaves only)
        domain_classif_slugs = []
        for _domain_sid, _p, _t in scopes_raw:
            if _p == "*" and _t:  # domain-level scope with tab
                _slug = _domain_sid.lower().replace(":", "_")
                _fname = f"classification_template_{_slug}.txt"
                _content = build_scope_classification_template(scope_nodes, domain_scope_id=_domain_sid)
                domain_classif_slugs.append(_fname)
                if not dry_run:
                    (PROMPTS_DIR / _fname).write_text(_content)

        # Combined fallback (all domains, backward compat for any code that still loads scope_classification_template.txt)
        _combined = build_scope_classification_template(scope_nodes, domain_scope_id=None)
        if not dry_run:
            (PROMPTS_DIR / "scope_classification_template.txt").write_text(_combined)

        (PROMPTS_DIR / "vehicle_type_template.txt").write_text(vehicle_type_template)
        (PROMPTS_DIR / "extraction_template.txt").write_text(extraction_template)
        for _sheet, _tmpl in type_templates.items():
            _slug = _sheet.lower().replace(" ", "_")
            (PROMPTS_DIR / f"extraction_template_{_slug}.txt").write_text(_tmpl)
        (PROMPTS_DIR / "extraction_retry_system.txt").write_text(
            "You are a data extraction assistant. Extract facts from tender documents into JSON. Output ONLY valid JSON. No markdown fences, no explanations.")
        (PROMPTS_DIR / "extraction_retry_template.txt").write_text(retry_template)

        # Prune stale extraction_template_*.txt files not in the current vt_prompt_map
        expected = set(vt_prompt_map.values()) | {
            "extraction_template.txt", "extraction_retry_template.txt",
            "extraction_retry_system.txt", "extraction_system.txt",
            "vehicle_type_template.txt", "scope_classification_template.txt",
            "basic_system.txt", "basic_template.txt", "nace_template.txt",
            "domain_detection_template.txt",
        }
        for _f in PROMPTS_DIR.glob("classification_template_*.txt"):
            expected.add(_f.name)
        for stale in PROMPTS_DIR.glob("extraction_template_*.txt"):
            if stale.name not in expected:
                stale.unlink()
                print(f"  [PRUNE] Deleted stale prompt: {stale.name}")
        (PROMPTS_DIR / "basic_system.txt").write_text(
            "You are a data extraction assistant. Extract facts from tender documents into JSON. Output ONLY valid JSON. No markdown fences, no explanations.")
        (PROMPTS_DIR / "basic_template.txt").write_text(basic_template)
        (PROMPTS_DIR / "contact_system.txt").write_text(
            "You are a contact-data extraction assistant. Extract contact details from the document excerpt. Output ONLY valid JSON. No markdown, no explanation.")
        _contact_skeleton = "{" + ",".join(f'"{k}":null' for k in _contact_target_keys) + "}"
        (PROMPTS_DIR / "contact_template.txt").write_text(
            "Extract contact details from this document excerpt. Use null if not found.\n\n"
            "DOCUMENT EXCERPT (last pages):\n{text}\n\n"
            f"Output ONLY this JSON:\n{_contact_skeleton}")
        (PROMPTS_DIR / "nace_system.txt").write_text(
            "You are an industrial classification specialist. Pick the single best NACE code from the provided list. Output ONLY valid JSON with exactly the field names shown.")
        (PROMPTS_DIR / "nace_template.txt").write_text(nace_template)
        (PROMPTS_DIR / "domain_detection_template.txt").write_text(domain_detection_template)

        # Sync per-domain READMEs from Spec/ → config/ (one file per domain)
        for _domain_sid, _parent, _tab in scopes_raw:
            if _parent != "*":
                continue
            _slug = _domain_sid.lower().replace(":", "_")
            _readme_spec  = ROOT / "Spec" / f"haystacked_industry_readme_{_slug}.md"
            _readme_local = CONFIG_DIR / f"industry_readme_{_slug}.md"
            if not _readme_spec.exists():
                raise SystemExit(
                    f"Missing domain README for {_domain_sid}: {_readme_spec}"
                )
            _readme_local.write_bytes(_readme_spec.read_bytes())
            print(f"  README synced: industry_readme_{_slug}.md")
        # Remove legacy single-domain file to prevent silent fallback
        _legacy_readme = CONFIG_DIR / "industry_readme.md"
        if _legacy_readme.exists():
            _legacy_readme.unlink()
            print("  Removed legacy config/industry_readme.md")

        (CONFIG_DIR / "ap0_checksum.txt").write_text(xlsx_md5)

        emit_fields_json(wb, data_sheets, plausibility_raw=plausibility_raw, tab_scope_map=tab_scope_map, scope_nodes=scope_nodes)

        print(f"\nWrote all config files. AP0 checksum: {xlsx_md5[:8]}…")

    if warnings:
        print(f"\nCONSISTENCY WARNINGS ({len(warnings)}):")
        for w in warnings: print(f"  [WARN] {w}")
        return 1
    print("\nAll consistency checks passed.")
    return 0


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate all runtime config from AP0 xlsx")
    parser.add_argument("--xlsx",     default=str(DEFAULT_XLSX))
    parser.add_argument("--platform", default=str(DEFAULT_PLATFORM))
    parser.add_argument("--db",       default=str(DEFAULT_DB))
    parser.add_argument("--dry-run",  action="store_true")
    args = parser.parse_args()
    sys.exit(generate(Path(args.xlsx), Path(args.db), args.dry_run, Path(args.platform)))
