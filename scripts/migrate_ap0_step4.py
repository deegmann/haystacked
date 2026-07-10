#!/usr/bin/env python3
"""One-shot AP0 Step 4 migration: Global tab + field movement + tab renames."""

import shutil
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import sys
    sys.exit("openpyxl not installed: pip3 install openpyxl")

ROOT   = Path(__file__).parent.parent
XLSX   = ROOT / "Spec" / "haystacked_AP0_field_spec_v0_10.xlsx"
BACKUP = ROOT / "Spec" / "haystacked_AP0_field_spec_v0_10_pre_step4_backup.xlsx"

GLOBAL_FIELDS = {
    "country", "employee_count_range", "founding_year", "hq_city",
    "certifications_generic", "languages_spoken",
    "reference_count", "lead_time_weeks", "distribution_model", "service_coverage",
}

# ── A1. Backup ────────────────────────────────────────────────────────────────
shutil.copy2(XLSX, BACKUP)
print(f"Backup: {BACKUP.name}")

# ── A2. Open workbook (NOT read_only) ─────────────────────────────────────────
wb = openpyxl.load_workbook(str(XLSX))

# ── A3. Identify column structure in SHARED tab ────────────────────────────────
shared_ws = wb["SHARED – All AGV Types"]
header_row_idx = None
n_cols = shared_ws.max_column

for row_idx in range(1, shared_ws.max_row + 1):
    cell_val = shared_ws.cell(row_idx, 1).value
    if cell_val == "Field Name":
        header_row_idx = row_idx
        break

if header_row_idx is None:
    raise ValueError("Header 'Field Name' not found in SHARED tab")

print(f"SHARED: header at row {header_row_idx}, {n_cols} columns")

data_rows_before = sum(
    1 for r in range(header_row_idx + 1, shared_ws.max_row + 1)
    if shared_ws.cell(r, 1).value
    and not str(shared_ws.cell(r, 1).value).startswith("──")
)
print(f"SHARED data rows before: {data_rows_before}")

# ── A4. Create Global tab at position 0 ───────────────────────────────────────
global_ws = wb.create_sheet("Global", 0)

# Copy header row from SHARED to Global row 1
for col_idx in range(1, n_cols + 1):
    global_ws.cell(1, col_idx, shared_ws.cell(header_row_idx, col_idx).value)

# ── A5. Move 10 field rows from SHARED to Global ──────────────────────────────
rows_to_move = []  # [(shared_row_idx, [values...]), ...]
for row_idx in range(header_row_idx + 1, shared_ws.max_row + 1):
    fname = shared_ws.cell(row_idx, 1).value
    if fname and str(fname) in GLOBAL_FIELDS:
        row_vals = [shared_ws.cell(row_idx, c).value for c in range(1, n_cols + 1)]
        rows_to_move.append((row_idx, row_vals))

print(f"Fields to move: {[r[1][0] for r in rows_to_move]}")

# Append to Global tab (starting at row 2)
for dest_idx, (_, row_vals) in enumerate(rows_to_move, 2):
    for col_idx, val in enumerate(row_vals, 1):
        global_ws.cell(dest_idx, col_idx, val)

# Delete from SHARED in reverse order (bottom-to-top avoids index shifting)
for row_idx, _ in sorted(rows_to_move, key=lambda x: x[0], reverse=True):
    shared_ws.delete_rows(row_idx)
    print(f"  Deleted SHARED row {row_idx}")

# ── A6. Update ③ Scope Registry tab_names ────────────────────────────────────
scope_ws = wb["③ Scope Registry"]
scope_header_row = None
scope_id_col = None
tab_name_col  = None

for row_idx in range(1, scope_ws.max_row + 1):
    cell_val = scope_ws.cell(row_idx, 1).value
    if cell_val == "scope_id":
        scope_header_row = row_idx
        for col_idx in range(1, scope_ws.max_column + 1):
            hval = scope_ws.cell(row_idx, col_idx).value
            if hval == "scope_id":
                scope_id_col = col_idx
            elif hval == "tab_name":
                tab_name_col = col_idx
        break

if scope_header_row is None or tab_name_col is None:
    raise ValueError("③ Scope Registry header/tab_name column not found")

SCOPE_TAB_UPDATES = {
    "*":                       "Global",
    "Logistics:AGV":           "AGV_Shared",
    "Logistics:AGV:Forklift":  "AGV_Forklift",
    "Logistics:AGV:Tugger":    "AGV_Tugger",
    "Logistics:AGV:AMR":       "AGV_AMR",
}

for row_idx in range(scope_header_row + 1, scope_ws.max_row + 1):
    sid = scope_ws.cell(row_idx, scope_id_col).value
    if sid and str(sid) in SCOPE_TAB_UPDATES:
        new_tab = SCOPE_TAB_UPDATES[str(sid)]
        scope_ws.cell(row_idx, tab_name_col, new_tab)
        print(f"  Scope Registry: {sid} → tab_name={new_tab}")

# ── A7. Rename physical tabs ──────────────────────────────────────────────────
wb["SHARED – All AGV Types"].title = "AGV_Shared"
wb["Forklift AGV"].title           = "AGV_Forklift"
wb["Tugger AGV"].title             = "AGV_Tugger"
wb["Mobile AMR"].title             = "AGV_AMR"

# ── A8. Save ──────────────────────────────────────────────────────────────────
wb.save(str(XLSX))
print("Step 4 migration complete.")

# ── Verification ──────────────────────────────────────────────────────────────
wb2 = openpyxl.load_workbook(str(XLSX), read_only=True, data_only=True)
print("\nVerification:")
print("  Physical tabs:", wb2.sheetnames)

global_data = sum(
    1 for r in wb2["Global"].iter_rows(min_row=2, values_only=True)
    if r and r[0] and not str(r[0]).startswith("──")
)
print(f"  Global data rows: {global_data} (expected: {len(rows_to_move)})")

shared_data_after = sum(
    1 for r in wb2["AGV_Shared"].iter_rows(values_only=True)
    if r and r[0]
    and r[0] != "Field Name"
    and not str(r[0]).startswith("──")
    and "AGV/AMR subtypes" not in str(r[0])  # skip title row
)
print(f"  AGV_Shared data rows: {shared_data_after} (expected: {data_rows_before - len(rows_to_move)})")
