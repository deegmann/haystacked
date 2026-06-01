"""
Schritt 1 — Einmalige Migration: Fügt Scoring Rule, Threshold 1, Threshold 2
als Spalten in die AP0 xlsx ein, direkt nach Scoring Weight.

Ausführen: python3 scripts/migrate_ap0_scoring_rules.py
Danach: python3 scripts/generate_all.py
"""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed: pip3 install openpyxl")

AP0 = Path(__file__).parent.parent / "Spec" / "haystacked_AP0_field_spec_v0_10.xlsx"

# Scoring rules per sheet and field name.
# Format: field → (rule, threshold_1, threshold_2)
RULES: dict[str, dict[str, tuple]] = {
    "SHARED – All AGV Types": {
        "stop_accuracy_mm":    ("tiered_lower",    10,  30),
        "battery_runtime_h":   ("threshold_upper",  8,  None),
        "autonomous_charging": ("bool",           None, None),
        "safety_standard":     ("nonempty",       None, None),
        "vda5050_compatible":  ("bool_cond",      None, None),
        "reference_count":     ("proportional",   20,  None),
        "lead_time_weeks":     ("tiered_lower",   26,   52),
    },
    "Forklift AGV": {
        "drop_accuracy_lat_mm":     ("threshold_lower", 20, None),
        "pick_req_accuracy_lat_mm": ("threshold_lower", 20, None),
        "stacking_capability":      ("bool",           None, None),
    },
    "Tugger AGV": {
        "auto_hitch":                  ("bool",    None, None),
        "trailer_steering_technology": ("nonempty", None, None),
    },
    "Mobile AMR": {
        "rotation_capable":          ("bool",        None, None),
        "throughput_picks_per_hour": ("tiered_upper", 300, 150),
    },
}


def migrate() -> int:
    if not AP0.exists():
        print(f"[ERROR] AP0 xlsx not found: {AP0}")
        return 1

    print(f"Opening: {AP0.name}")
    wb = openpyxl.load_workbook(str(AP0))
    total = 0

    for sheet_name, rules in RULES.items():
        if sheet_name not in wb.sheetnames:
            print(f"  [SKIP] Sheet not found: {sheet_name!r}")
            continue

        ws = wb[sheet_name]

        # Find header row (first cell = "Field Name")
        hrow = None
        for row in ws.iter_rows(max_row=30):
            if row[0].value == "Field Name":
                hrow = row[0].row
                break
        if hrow is None:
            print(f"  [SKIP] No header row in {sheet_name!r}")
            continue

        # Build col-name → col-index map from header row
        hcols = {
            ws.cell(hrow, c).value: c
            for c in range(1, ws.max_column + 1)
            if ws.cell(hrow, c).value
        }

        if "Scoring Weight" not in hcols:
            print(f"  [SKIP] No 'Scoring Weight' column in {sheet_name!r}")
            continue

        # Assign columns for new headers (append after last used column)
        max_col = ws.max_column
        if "Scoring Rule" not in hcols:
            rule_col = max_col + 1
            t1_col   = max_col + 2
            t2_col   = max_col + 3
            ws.cell(hrow, rule_col).value = "Scoring Rule"
            ws.cell(hrow, t1_col).value   = "Threshold 1"
            ws.cell(hrow, t2_col).value   = "Threshold 2"
        else:
            rule_col = hcols["Scoring Rule"]
            t1_col   = hcols.get("Threshold 1", max_col + 1)
            t2_col   = hcols.get("Threshold 2", max_col + 2)
            if "Threshold 1" not in hcols:
                ws.cell(hrow, t1_col).value = "Threshold 1"
            if "Threshold 2" not in hcols:
                ws.cell(hrow, t2_col).value = "Threshold 2"

        fn_col = hcols.get("Field Name", 1)
        count = 0
        for row in ws.iter_rows(min_row=hrow + 1):
            cell_fn = row[fn_col - 1]
            if not cell_fn.value:
                continue
            fname = str(cell_fn.value).strip()
            if fname.startswith("──") or fname not in rules:
                continue
            rule, t1, t2 = rules[fname]
            ws.cell(cell_fn.row, rule_col).value = rule
            ws.cell(cell_fn.row, t1_col).value   = t1
            ws.cell(cell_fn.row, t2_col).value   = t2
            print(f"    {sheet_name}/{fname}: {rule}, t1={t1}, t2={t2}")
            count += 1

        print(f"  → {sheet_name}: {count}/{len(rules)} fields updated")
        total += count

    wb.save(str(AP0))
    print(f"\n{total} fields written. Saved: {AP0.name}")
    print("Next: python3 scripts/generate_all.py")
    return 0


if __name__ == "__main__":
    sys.exit(migrate())
