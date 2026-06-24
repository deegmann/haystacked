#!/usr/bin/env python3
"""Add empty 'Client Explanation' column after 'Display Mode' in all AP0 DATA_SHEETS."""
from pathlib import Path
import openpyxl

ROOT      = Path(__file__).parent.parent
XLSX_PATH = ROOT / "Spec" / "haystacked_AP0_field_spec_v0_10.xlsx"
DATA_SHEETS = ["SHARED – All AGV Types", "Forklift AGV", "Tugger AGV", "Mobile AMR"]
AFTER_HEADER = "Display Mode"
NEW_HEADER   = "Client Explanation"

wb = openpyxl.load_workbook(XLSX_PATH)  # NO read_only, NO data_only

for sheet_name in DATA_SHEETS:
    if sheet_name not in wb.sheetnames:
        print(f"  [SKIP] Sheet not found: {sheet_name}")
        continue
    ws = wb[sheet_name]

    # Find header row — first row with 3+ non-empty cells
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if sum(1 for c in row if c and str(c).strip()) >= 3:
            header_row_idx = i
            break
    if header_row_idx is None:
        print(f"  [WARN] No header row in {sheet_name}")
        continue

    # Check if column already exists
    header_cells = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]
    if NEW_HEADER in [str(c).strip() for c in header_cells if c]:
        print(f"  [SKIP] '{NEW_HEADER}' already exists in {sheet_name}")
        continue

    # Find insertion point (after AFTER_HEADER)
    insert_after = next(
        (i + 1 for i, h in enumerate(header_cells) if h and str(h).strip() == AFTER_HEADER),
        None
    )
    if insert_after is None:
        print(f"  [WARN] '{AFTER_HEADER}' column not found in {sheet_name} — appending at end")
        insert_after = len([c for c in header_cells if c])

    insert_col = insert_after + 1  # 1-based, insert AFTER the found column
    ws.insert_cols(insert_col)
    ws.cell(row=header_row_idx, column=insert_col, value=NEW_HEADER)
    print(f"  Added '{NEW_HEADER}' at col {insert_col} in {sheet_name}")

wb.save(XLSX_PATH)
print(f"\nSaved {XLSX_PATH}")
