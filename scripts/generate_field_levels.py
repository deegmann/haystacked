"""
generate_field_levels.py
Single source of truth: AP0 xlsx → config/field_levels.json

The AP0 xlsx is the only authoritative source for field classifications AND
matching operators. No overrides exist here — all changes must go into the
xlsx first.

field_levels.json output format (v0.8+):
  {
    "max_payload_kg": { "level": "KO", "operator": "KO_IF_LT" },
    "vna_capable":    { "level": "COND_KO", "operator": "KO_BOOL_EXCLUSIVE" },
    "reference_count":{ "level": "SCORING" },
    ...
  }

Matching operators (defined in AP0 xlsx "Matching Operator" column):
  KO_IF_LT         — K.O. if supplier value < tender value
  KO_IF_GT         — K.O. if supplier value > tender value
  KO_IF_NEQ        — K.O. if supplier value ≠ tender value
  KO_BOOL_REQUIRED — K.O. if tender=required and supplier≠True
  KO_BOOL_EXCLUSIVE— Bidirectional: required→supplier must=True; not_required→supplier must≠True
  KO_SUBSET        — K.O. if no overlap between tender list and supplier list

Also validates consistency between AP0 xlsx and the SQLite schema.

Usage:
    python3 scripts/generate_field_levels.py [--xlsx PATH] [--db PATH] [--dry-run]
"""

import argparse
import json
import sqlite3
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed: pip3 install openpyxl")

ROOT         = Path(__file__).parent.parent
DEFAULT_XLSX = Path("/Users/christiandeeg/haystacked_platform/Specs/haystacked_AP0_field_spec_v0_10.xlsx")
DEFAULT_DB   = ROOT / "data" / "haystacked.db"
OUTPUT_JSON  = ROOT / "config" / "field_levels.json"

XLSX_LEVEL_MAP = {
    "K.O.":       "KO",
    "Cond. K.O.": "COND_KO",
    "Scoring":    "SCORING",
    "Context":    "CONTEXT",
}

VALID_OPERATORS = {
    "KO_IF_LT", "KO_IF_GT", "KO_IF_NEQ",
    "KO_BOOL_REQUIRED", "KO_BOOL_EXCLUSIVE", "KO_SUBSET",
}

XLSX_SHEETS = [
    "SHARED – All AGV Types",
    "Forklift AGV",
    "Tugger AGV",
    "Mobile AMR",
]

# SQLite columns that are structural PKs/FKs or duplicates — not matching fields.
SQLITE_STRUCTURAL = {
    "extension_id", "base_model_id", "company_id", "product_id",
    "company_name", "product_name", "product_description",
    "base_model_name", "oem_company_id", "oem_link_public",
    "website", "last_updated", "export_capable", "is_oem_product", "active",
    "extra_fields",
    "pick_req_accuracy_lat_mm_amr", "pick_req_accuracy_dep_mm_amr",
    "pick_req_accuracy_angle_deg_amr",
    "drop_accuracy_lat_mm_amr",  "drop_accuracy_dep_mm_amr",
    "drop_accuracy_angle_deg_amr",
    "min_project_value_eur", "max_project_value_eur",
}


def read_xlsx(xlsx_path: Path) -> dict[str, dict]:
    """Read field → {level, operator, tender_key, data_type} from all sheets in AP0 xlsx."""
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    fields: dict[str, dict] = {}
    missing_sheets = []

    for sheet_name in XLSX_SHEETS:
        if sheet_name not in wb.sheetnames:
            missing_sheets.append(sheet_name)
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        col_field = col_level = col_operator = col_tender_key = col_data_type = None
        header_idx = None
        for i, row in enumerate(rows):
            if row and row[0] == "Field Name":
                header_idx = i
                for j, cell in enumerate(row):
                    if cell == "Field Name":       col_field = j
                    if cell == "Level":            col_level = j
                    if cell == "Matching Operator": col_operator = j
                    if cell == "Tender JSON Key":  col_tender_key = j
                    if cell == "Data Type":        col_data_type = j
                break

        if header_idx is None or col_level is None:
            missing_sheets.append(f"{sheet_name} (no header row found)")
            continue

        if col_operator is None:
            print(f"  [WARN] Sheet '{sheet_name}' has no 'Matching Operator' column — operators will be absent")

        for row in rows[header_idx + 1:]:
            if not row or not row[col_field]:
                continue
            fname = str(row[col_field]).strip()
            raw_level = row[col_level]
            if not fname or not raw_level or fname.startswith("──"):
                continue
            level = XLSX_LEVEL_MAP.get(str(raw_level).strip())
            if not level:
                continue

            entry: dict = {"level": level}

            if col_data_type is not None and col_data_type < len(row) and row[col_data_type]:
                entry["data_type"] = str(row[col_data_type]).strip()

            if col_operator is not None and col_operator < len(row):
                raw_op = row[col_operator]
                if raw_op:
                    op = str(raw_op).strip()
                    if op in VALID_OPERATORS:
                        entry["operator"] = op
                    else:
                        print(f"  [WARN] '{fname}': unknown operator '{op}' — ignored")

            if col_tender_key is not None and col_tender_key < len(row) and row[col_tender_key]:
                entry["tender_key"] = str(row[col_tender_key]).strip()

            # Warn if KO/COND_KO field has no operator
            if level in ("KO", "COND_KO") and "operator" not in entry:
                print(f"  [WARN] '{fname}' is {level} but has no Matching Operator defined")

            # Warn if KO/COND_KO field with operator has no tender_key
            if level in ("KO", "COND_KO") and "operator" in entry and "tender_key" not in entry:
                print(f"  [WARN] '{fname}' has operator but no Tender JSON Key — cannot match tender requirements")

            if fname in fields and fields[fname] != entry:
                print(f"  [WARN] '{fname}' defined in multiple sheets "
                      f"({fields[fname]} vs {entry}) — last occurrence wins ({sheet_name})")
            fields[fname] = entry

    if missing_sheets:
        print(f"  [WARN] Sheets not found in xlsx: {missing_sheets}")

    return fields


def read_sqlite_columns(db_path: Path) -> dict[str, list[str]]:
    """Return {table: [col, ...]} for the three matching tables."""
    if not db_path.exists():
        return {}
    conn = sqlite3.connect(str(db_path))
    result = {}
    for table in ("companies", "products", "base_model_extensions"):
        cols = [r[1] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()]
        result[table] = cols
    conn.close()
    return result


def validate(xlsx_fields: dict[str, dict],
             sqlite_cols: dict[str, list[str]]) -> list[str]:
    """Compare AP0 xlsx fields against SQLite schema."""
    warnings: list[str] = []

    if not sqlite_cols:
        warnings.append("SQLite DB not found — schema validation skipped")
        return warnings

    all_db_fields: set[str] = set()
    for cols in sqlite_cols.values():
        all_db_fields.update(cols)
    all_db_fields -= SQLITE_STRUCTURAL

    xlsx_set = set(xlsx_fields.keys())

    for f in sorted(xlsx_set - all_db_fields):
        warnings.append(
            f"AP0 xlsx defines '{f}' ({xlsx_fields[f]['level']}) "
            f"but the field is missing from SQLite — data not collected in Airtable"
        )
    for f in sorted(all_db_fields - xlsx_set):
        warnings.append(
            f"SQLite has field '{f}' with no classification in AP0 xlsx "
            f"— unclassified, excluded from matching logic"
        )

    return warnings


def generate(xlsx_path: Path, db_path: Path, dry_run: bool = False) -> int:
    """Read AP0 xlsx, validate against SQLite, write field_levels.json.
    Returns 0 if fully consistent, 1 if warnings exist.
    """
    print(f"Reading AP0 xlsx: {xlsx_path.name}")
    xlsx_fields = read_xlsx(xlsx_path)
    ko_count   = sum(1 for v in xlsx_fields.values() if v["level"] == "KO")
    cond_count = sum(1 for v in xlsx_fields.values() if v["level"] == "COND_KO")
    op_count   = sum(1 for v in xlsx_fields.values() if "operator" in v)
    key_count  = sum(1 for v in xlsx_fields.values() if "tender_key" in v)
    print(f"  {len(xlsx_fields)} fields read  ({ko_count} KO, {cond_count} COND_KO, {op_count} with operator, {key_count} with tender_key)")

    print(f"Reading SQLite schema: {db_path}")
    sqlite_cols = read_sqlite_columns(db_path)
    if sqlite_cols:
        total = sum(len(v) for v in sqlite_cols.values())
        print(f"  {total} columns across 3 tables")
    else:
        print("  [WARN] DB not found — schema validation skipped")

    warnings = validate(xlsx_fields, sqlite_cols)

    if dry_run:
        print("\n[DRY RUN] Would write field_levels.json:")
        print(json.dumps(xlsx_fields, indent=2))
    else:
        OUTPUT_JSON.write_text(json.dumps(xlsx_fields, indent=2, ensure_ascii=False) + "\n")
        print(f"\nWrote {OUTPUT_JSON} ({len(xlsx_fields)} fields)")

    if warnings:
        print(f"\n{'='*60}")
        print(f"CONSISTENCY WARNINGS ({len(warnings)}):")
        for w in warnings:
            print(f"  [WARN] {w}")
        print(f"{'='*60}")
        return 1

    print("\nAll consistency checks passed.")
    return 0


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generate field_levels.json from AP0 xlsx (single source of truth)"
    )
    parser.add_argument("--xlsx",    default=str(DEFAULT_XLSX), help="Path to AP0 field spec xlsx")
    parser.add_argument("--db",      default=str(DEFAULT_DB),   help="Path to haystacked.db")
    parser.add_argument("--dry-run", action="store_true",        help="Print output, do not write")
    args = parser.parse_args()

    sys.exit(generate(Path(args.xlsx), Path(args.db), args.dry_run))
