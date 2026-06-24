#!/usr/bin/env python3
"""Remove 'New', 'Tender Unit', 'Tender JSON Key', and 'Mand.' columns from all AP0 DATA_SHEETS."""
import sys
from pathlib import Path
import openpyxl

ROOT      = Path(__file__).parent.parent
XLSX_PATH = ROOT / "Spec" / "haystacked_AP0_field_spec_v0_10.xlsx"
DATA_SHEETS = ["SHARED – All AGV Types", "Forklift AGV", "Tugger AGV", "Mobile AMR"]
COLS_TO_REMOVE = {"New", "Tender Unit", "Tender JSON Key", "Mand."}

wb = openpyxl.load_workbook(XLSX_PATH)  # NO read_only, NO data_only

total_removed = 0
for sheet_name in DATA_SHEETS:
    if sheet_name not in wb.sheetnames:
        print(f"  [SKIP] Sheet not found: {sheet_name}")
        continue
    ws = wb[sheet_name]

    # Find header row — first row where at least 3 cells have non-empty string values
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        non_empty = sum(1 for c in row if c and str(c).strip())
        if non_empty >= 3:
            header_row_idx = i
            break
    if header_row_idx is None:
        print(f"  [WARN] No header row found in {sheet_name}")
        continue

    # Find columns to delete (collect indices in reverse order so deletion doesn't shift)
    header_cells = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]
    cols_to_delete = sorted(
        [i + 1 for i, h in enumerate(header_cells) if h and str(h).strip() in COLS_TO_REMOVE],
        reverse=True  # delete right-to-left to preserve indices
    )

    for col_idx in cols_to_delete:
        col_name = ws.cell(row=header_row_idx, column=col_idx).value
        ws.delete_cols(col_idx)
        print(f"  Removed column '{col_name}' (was col {col_idx}) from {sheet_name}")
        total_removed += 1

if total_removed == 0:
    print("No columns removed — already clean or column names not found.")
else:
    wb.save(XLSX_PATH)
    print(f"\nSaved {XLSX_PATH}. Total columns removed: {total_removed}")
